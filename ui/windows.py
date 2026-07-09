import datetime
import os
import platform
import re
import subprocess
from tkinter import filedialog, messagebox

import customtkinter as ctk
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from database import Database, ServiceItem
from ui.theme import PALETTE, button_style, card


def styled_button(master, text: str, command, kind: str = "primary", width: int = 120):
    return ctk.CTkButton(master, text=text, command=command, width=width, **button_style(kind))


def normalize_rf_phone(phone: str) -> str | None:
    digits = re.sub(r"\D", "", phone or "")
    if len(digits) == 11 and digits[0] in ("7", "8"):
        digits = "7" + digits[1:]
    elif len(digits) == 10 and digits[0] == "9":
        digits = "7" + digits
    else:
        return None
    return f"+{digits}"


def format_rf_phone(phone: str) -> str:
    normalized = normalize_rf_phone(phone)
    if not normalized:
        return phone
    digits = normalized[1:]
    return f"+7 {digits[1:4]} {digits[4:7]} {digits[7:9]} {digits[9:11]}"


class ServiceManager(ctk.CTkToplevel):
    def __init__(self, parent, db: Database, on_change):
        super().__init__(parent)
        self.db = db
        self.on_change = on_change
        self.title("Управление услугами")
        self.geometry("980x620")
        self.minsize(860, 520)
        self.configure(fg_color=PALETTE["bg"])
        self.setup_ui()
        self.load_services()

    def setup_ui(self):
        top = card(self)
        top.pack(fill="x", padx=14, pady=(14, 8))
        ctk.CTkLabel(top, text="Каталог услуг", font=("Arial", 20, "bold")).pack(side="left", padx=12, pady=10)
        actions = ctk.CTkFrame(top, fg_color="transparent")
        actions.pack(side="right", padx=10)
        styled_button(actions, "Добавить", self.add_service_dialog, "success", 120).pack(side="left", padx=4, pady=10)
        styled_button(actions, "Обновить", self.load_services, "primary", 120).pack(side="left", padx=4, pady=10)
        styled_button(actions, "Закрыть", self.destroy, "ghost", 120).pack(side="left", padx=4, pady=10)

        self.table_wrap = card(self)
        self.table_wrap.pack(fill="both", expand=True, padx=14, pady=(0, 14))
        self.table = ctk.CTkScrollableFrame(self.table_wrap, height=480, fg_color="transparent")
        self.table.pack(fill="both", expand=True, padx=10, pady=10)

    def load_services(self):
        for child in self.table.winfo_children():
            child.destroy()
        header = ctk.CTkFrame(self.table, fg_color=PALETTE["card_alt"], corner_radius=8)
        header.pack(fill="x", pady=(0, 4))
        ctk.CTkLabel(header, text="Название", width=370, anchor="w", font=("Arial", 13, "bold")).pack(side="left", padx=8, pady=8)
        ctk.CTkLabel(header, text="Цена", width=120, anchor="e", font=("Arial", 13, "bold")).pack(side="left", padx=8)
        ctk.CTkLabel(header, text="Категория", width=170, anchor="w", font=("Arial", 13, "bold")).pack(side="left", padx=8)
        ctk.CTkLabel(header, text="Статус", width=90, anchor="center", font=("Arial", 13, "bold")).pack(side="left", padx=8)
        ctk.CTkLabel(header, text="Действия", width=230, anchor="center", font=("Arial", 13, "bold")).pack(side="left", padx=8)

        for row in self.db.get_all_services():
            active = bool(row["is_active"])
            row_frame = ctk.CTkFrame(
                self.table,
                fg_color=("#ECFDF5", "#10261A") if active else ("#FFF7ED", "#2D1F12"),
                corner_radius=8,
            )
            row_frame.pack(fill="x", pady=2)
            ctk.CTkLabel(row_frame, text=row["name"], width=370, anchor="w").pack(side="left", padx=8, pady=6)
            ctk.CTkLabel(row_frame, text=f"{row['price']:.2f}", width=120, anchor="e").pack(side="left", padx=8)
            ctk.CTkLabel(row_frame, text=row["category"], width=170, anchor="w").pack(side="left", padx=8)
            ctk.CTkLabel(row_frame, text="Активна" if active else "Отключена", width=90, anchor="center").pack(side="left", padx=8)
            styled_button(row_frame, "Редакт.", lambda sid=row["id"]: self.edit_service_dialog(sid), "primary", 72).pack(side="left", padx=3)
            styled_button(
                row_frame,
                "Откл." if active else "Вкл.",
                lambda sid=row["id"], st=active: self.toggle_service(sid, st),
                "warning",
                64,
            ).pack(side="left", padx=3)
            styled_button(row_frame, "Удал.", lambda sid=row["id"]: self.delete_service(sid), "danger", 64).pack(side="left", padx=3)

    def _service_dialog(self, title: str, service=None):
        result: dict[str, str | float | None] = {"name": None, "price": None, "category": None}
        dialog = ctk.CTkToplevel(self)
        dialog.title(title)
        dialog.geometry("520x340")
        dialog.resizable(False, False)
        dialog.configure(fg_color=PALETTE["bg"])
        dialog.grab_set()

        content = card(dialog)
        content.pack(fill="both", expand=True, padx=12, pady=12)
        ctk.CTkLabel(content, text=title, font=("Arial", 18, "bold")).pack(anchor="w", padx=12, pady=(12, 6))

        ctk.CTkLabel(content, text="Название услуги").pack(anchor="w", padx=12)
        name_entry = ctk.CTkEntry(content, corner_radius=10, width=460)
        name_entry.pack(padx=12, pady=(2, 8))
        name_entry.insert(0, service["name"] if service else "")

        ctk.CTkLabel(content, text="Цена").pack(anchor="w", padx=12)
        price_entry = ctk.CTkEntry(content, corner_radius=10, width=460)
        price_entry.pack(padx=12, pady=(2, 8))
        price_entry.insert(0, f"{float(service['price']):.2f}" if service else "")

        ctk.CTkLabel(content, text="Категория").pack(anchor="w", padx=12)
        categories = self.db.get_categories()
        if not categories:
            categories = ["Основные"]
        current_category = service["category"] if service else categories[0]
        if current_category not in categories:
            categories.append(current_category)
        category_var = ctk.StringVar(value=current_category)
        category_menu = ctk.CTkOptionMenu(content, values=categories, variable=category_var, width=330)
        category_menu.pack(anchor="w", padx=12, pady=(2, 8))

        def add_new_category():
            category_dialog = ctk.CTkInputDialog(text="Название новой категории", title="Новая категория")
            new_cat = (category_dialog.get_input() or "").strip()
            if not new_cat:
                return
            if new_cat not in categories:
                categories.append(new_cat)
                category_menu.configure(values=categories)
            category_var.set(new_cat)

        styled_button(content, "Добавить категорию", add_new_category, "secondary", 170).pack(anchor="w", padx=12, pady=(0, 10))

        button_row = ctk.CTkFrame(content, fg_color="transparent")
        button_row.pack(fill="x", padx=12, pady=(0, 12))

        def submit():
            name = name_entry.get().strip()
            category = category_var.get().strip()
            try:
                price = float(price_entry.get().strip().replace(",", "."))
            except ValueError:
                messagebox.showwarning("Ошибка", "Введите корректную цену")
                return
            if not name or not category or price < 0:
                messagebox.showwarning("Ошибка", "Заполните все поля корректно")
                return
            result["name"] = name
            result["price"] = price
            result["category"] = category
            dialog.destroy()

        styled_button(button_row, "Сохранить", submit, "success", 130).pack(side="right", padx=6)
        styled_button(button_row, "Отмена", dialog.destroy, "ghost", 110).pack(side="right", padx=6)

        self.wait_window(dialog)
        if result["name"] is None:
            return None
        return str(result["name"]), float(result["price"]), str(result["category"])

    def add_service_dialog(self):
        data = self._service_dialog("Добавить услугу")
        if not data:
            return
        name, price, category = data
        if not self.db.add_service(name, price, category):
            messagebox.showwarning("Ошибка", "Услуга уже существует")
            return
        self.load_services()
        self.on_change()

    def edit_service_dialog(self, service_id: int):
        service = next((s for s in self.db.get_all_services() if s["id"] == service_id), None)
        if not service:
            return
        data = self._service_dialog("Редактировать услугу", service)
        if not data:
            return
        name, price, category = data
        if not self.db.update_service(service_id, name, price, category):
            messagebox.showwarning("Ошибка", "Не удалось сохранить изменения")
            return
        self.load_services()
        self.on_change()

    def toggle_service(self, service_id: int, current_active: bool):
        self.db.set_service_active(service_id, not current_active)
        self.load_services()
        self.on_change()

    def delete_service(self, service_id: int):
        if not messagebox.askyesno("Подтверждение", "Удалить услугу?"):
            return
        self.db.delete_service(service_id)
        self.load_services()
        self.on_change()


class ClientSelector(ctk.CTkToplevel):
    def __init__(self, parent, db: Database):
        super().__init__(parent)
        self.db = db
        self.selected_client_id: int | None = None
        self.title("Выбор клиента")
        self.geometry("700x560")
        self.minsize(620, 480)
        self.configure(fg_color=PALETTE["bg"])
        self.setup_ui()
        self.load_clients()

    def setup_ui(self):
        top = card(self)
        top.pack(fill="x", padx=14, pady=(14, 8))
        ctk.CTkLabel(top, text="Выбор клиента", font=("Arial", 20, "bold")).pack(side="left", padx=12, pady=10)
        self.search = ctk.CTkEntry(top, placeholder_text="Имя или телефон", width=320, corner_radius=10)
        self.search.pack(side="left", padx=8)
        self.search.bind("<KeyRelease>", lambda _: self.load_clients())
        styled_button(top, "Новый клиент", self.create_client, "success", 130).pack(side="right", padx=6, pady=10)
        styled_button(top, "Выбрать", self.confirm, "primary", 120).pack(side="right", padx=6, pady=10)

        wrap = card(self)
        wrap.pack(fill="both", expand=True, padx=14, pady=(0, 14))
        self.list_frame = ctk.CTkScrollableFrame(wrap, height=420, fg_color="transparent")
        self.list_frame.pack(fill="both", expand=True, padx=10, pady=10)

    def load_clients(self):
        for child in self.list_frame.winfo_children():
            child.destroy()
        query = self.search.get().strip()
        rows = self.db.search_clients(query) if query else self.db.get_all_clients()

        header = ctk.CTkFrame(self.list_frame, fg_color=PALETTE["card_alt"], corner_radius=8)
        header.pack(fill="x", pady=(0, 4))
        ctk.CTkLabel(header, text="Имя", width=150, anchor="w", font=("Arial", 13, "bold")).pack(side="left", padx=8, pady=8)
        ctk.CTkLabel(header, text="Телефон", width=150, anchor="w", font=("Arial", 13, "bold")).pack(side="left", padx=8)
        ctk.CTkLabel(header, text="Комментарий", width=170, anchor="w", font=("Arial", 13, "bold")).pack(side="left", padx=8)
        ctk.CTkLabel(header, text="Заказов", width=70, anchor="center", font=("Arial", 13, "bold")).pack(side="left", padx=8)
        ctk.CTkLabel(header, text="Потрачено", width=110, anchor="e", font=("Arial", 13, "bold")).pack(side="left", padx=8)

        for row in rows:
            frame = ctk.CTkFrame(self.list_frame, fg_color=PALETTE["card_alt"], corner_radius=8)
            frame.pack(fill="x", pady=2)
            ctk.CTkLabel(frame, text=row["name"], width=150, anchor="w").pack(side="left", padx=8, pady=6)
            ctk.CTkLabel(frame, text=format_rf_phone(row["phone"]), width=150, anchor="w").pack(side="left", padx=8)
            ctk.CTkLabel(frame, text=(row["client_comment"] or "")[:26], width=170, anchor="w").pack(side="left", padx=8)
            ctk.CTkLabel(frame, text=str(row["total_orders"]), width=70, anchor="center").pack(side="left", padx=8)
            ctk.CTkLabel(frame, text=f"{row['total_spent']:.2f}", width=110, anchor="e").pack(side="left", padx=8)
            styled_button(frame, "Выбрать", lambda cid=row["id"]: self.select(cid), "primary", 100).pack(side="right", padx=8)

    def create_client(self):
        dialog = ctk.CTkToplevel(self)
        dialog.title("Новый клиент")
        dialog.geometry("520x360")
        dialog.resizable(False, False)
        dialog.configure(fg_color=PALETTE["bg"])
        dialog.grab_set()

        content = card(dialog)
        content.pack(fill="both", expand=True, padx=12, pady=12)
        ctk.CTkLabel(content, text="Новый клиент", font=("Arial", 18, "bold")).pack(anchor="w", padx=12, pady=(12, 6))

        ctk.CTkLabel(content, text="Имя клиента").pack(anchor="w", padx=12)
        name_entry = ctk.CTkEntry(content, width=460, corner_radius=10)
        name_entry.pack(padx=12, pady=(2, 8))

        ctk.CTkLabel(content, text="Телефон (+7 962 550 7832)").pack(anchor="w", padx=12)
        phone_entry = ctk.CTkEntry(content, width=460, corner_radius=10)
        phone_entry.pack(padx=12, pady=(2, 8))

        ctk.CTkLabel(content, text="Комментарий").pack(anchor="w", padx=12)
        comment_entry = ctk.CTkTextbox(content, width=460, height=90, corner_radius=10)
        comment_entry.pack(padx=12, pady=(2, 10))

        def submit():
            name = name_entry.get().strip()
            raw_phone = phone_entry.get().strip()
            comment = comment_entry.get("1.0", "end").strip()
            normalized_phone = normalize_rf_phone(raw_phone)
            if not name:
                messagebox.showwarning("Ошибка", "Введите имя клиента")
                return
            if not normalized_phone:
                messagebox.showwarning("Ошибка", "Введите корректный номер РФ, например: +7 962 550 7832")
                return
            existing_by_phone = self.db.find_client_by_phone(normalized_phone)
            if existing_by_phone:
                messagebox.showinfo("Инфо", "Клиент с таким телефоном уже существует, будет выбран существующий")
                self.select(int(existing_by_phone["id"]))
                return
            existing = self.db.find_client_by_name_phone(name, normalized_phone)
            if existing:
                self.select(int(existing["id"]))
                return
            client_id = self.db.create_client(name, normalized_phone, comment)
            if client_id is None:
                messagebox.showwarning("Ошибка", "Не удалось создать клиента")
                return
            self.select(client_id)

        row = ctk.CTkFrame(content, fg_color="transparent")
        row.pack(fill="x", padx=12, pady=(0, 12))
        styled_button(row, "Сохранить", submit, "success", 120).pack(side="right", padx=6)
        styled_button(row, "Отмена", dialog.destroy, "ghost", 110).pack(side="right", padx=6)

    def select(self, client_id: int):
        self.selected_client_id = client_id
        self.confirm()

    def confirm(self):
        if self.selected_client_id is None:
            messagebox.showwarning("Внимание", "Клиент не выбран")
            return
        self.destroy()


class OrderEditor(ctk.CTkToplevel):
    def __init__(self, parent, db: Database, order_id: int | None = None, client_id: int | None = None):
        super().__init__(parent)
        self.db = db
        self.order_id = order_id
        self.client_id = client_id
        self.items: list[ServiceItem] = []
        self.service_rows = self.db.get_active_services()
        self.selected_service_name: str | None = None
        self.title("Заказ-наряд")
        self.geometry("1220x760")
        self.minsize(1020, 620)
        self.configure(fg_color=PALETTE["bg"])
        self.setup_ui()
        if self.order_id:
            self.load_existing_order()
        self.refresh_services_panel()
        self.refresh_items_panel()

    def setup_ui(self):
        top = card(self)
        top.pack(fill="x", padx=14, pady=(14, 8))
        self.title_label = ctk.CTkLabel(top, text="Новый заказ", font=("Arial", 20, "bold"))
        self.title_label.pack(side="left", padx=12, pady=10)
        self.client_label = ctk.CTkLabel(top, text="Клиент: не выбран", font=("Arial", 14), text_color=PALETTE["muted"])
        self.client_label.pack(side="left", padx=12, pady=10)

        body = ctk.CTkFrame(self, fg_color="transparent")
        body.pack(fill="both", expand=True, padx=14, pady=(0, 14))

        left_wrap = card(body)
        left_wrap.pack(side="left", fill="both", expand=True, padx=(0, 6))
        right_wrap = card(body)
        right_wrap.pack(side="left", fill="both", expand=True, padx=(6, 0))

        ctk.CTkLabel(left_wrap, text="Каталог услуг", font=("Arial", 16, "bold")).pack(anchor="w", padx=12, pady=(12, 6))
        search_frame = ctk.CTkFrame(left_wrap, fg_color="transparent")
        search_frame.pack(fill="x", padx=10, pady=(0, 8))
        self.search_service = ctk.CTkEntry(search_frame, placeholder_text="Поиск услуги", corner_radius=10)
        self.search_service.pack(side="left", fill="x", expand=True, padx=(0, 6))
        self.search_service.bind("<KeyRelease>", lambda _: self.refresh_services_panel())
        styled_button(search_frame, "Услуги", self.open_service_manager, "secondary", 100).pack(side="left")

        self.services_panel = ctk.CTkScrollableFrame(left_wrap, height=430, fg_color="transparent")
        self.services_panel.pack(fill="both", expand=True, padx=10, pady=(0, 8))

        add_frame = ctk.CTkFrame(left_wrap, fg_color="transparent")
        add_frame.pack(fill="x", padx=10, pady=(0, 12))
        ctk.CTkLabel(add_frame, text="Кол-во:", width=60).pack(side="left")
        self.qty = ctk.CTkEntry(add_frame, width=70, corner_radius=10)
        self.qty.insert(0, "1")
        self.qty.pack(side="left", padx=(0, 8))
        styled_button(add_frame, "Добавить услугу", self.add_selected_service, "success", 160).pack(side="left")

        ctk.CTkLabel(right_wrap, text="Услуги в заказе", font=("Arial", 16, "bold")).pack(anchor="w", padx=12, pady=(12, 6))
        self.items_panel = ctk.CTkScrollableFrame(right_wrap, height=430, fg_color="transparent")
        self.items_panel.pack(fill="both", expand=True, padx=10, pady=(0, 8))

        bottom = ctk.CTkFrame(right_wrap, fg_color=PALETTE["card_alt"], corner_radius=10)
        bottom.pack(fill="x", padx=10, pady=(0, 12))
        self.total_label = ctk.CTkLabel(bottom, text="ИТОГО: 0.00", font=("Arial", 20, "bold"), text_color=PALETTE["primary"])
        self.total_label.pack(side="left", padx=10, pady=10)
        styled_button(bottom, "Печать", self.print_order, "warning", 120).pack(side="right", padx=6, pady=10)
        styled_button(bottom, "Сохранить", self.save_order, "success", 120).pack(side="right", padx=6, pady=10)

    def load_existing_order(self):
        order = self.db.get_order_by_id(self.order_id)
        if not order:
            return
        self.client_id = order["client_id"]
        self.title_label.configure(text=f"Заказ {order['order_number']}")
        self.client_label.configure(text=f"Клиент: {order['client_name']} | {order['phone']}")
        rows = self.db.get_order_services(self.order_id)
        self.items = [ServiceItem(int(r["id"]), r["service_name"], float(r["price"]), int(r["quantity"])) for r in rows]

    def open_service_manager(self):
        manager = ServiceManager(self, self.db, self.reload_services)
        manager.grab_set()

    def reload_services(self):
        self.service_rows = self.db.get_active_services()
        self.refresh_services_panel()

    def refresh_services_panel(self):
        for child in self.services_panel.winfo_children():
            child.destroy()
        query = self.search_service.get().strip().lower()
        rows = self.service_rows if not query else [s for s in self.service_rows if query in s["name"].lower()]
        for row in rows:
            name = row["name"]
            btn = ctk.CTkButton(
                self.services_panel,
                text=f"{name} — {row['price']:.2f}",
                anchor="w",
                fg_color=PALETTE["primary"] if self.selected_service_name == name else PALETTE["card_alt"],
                hover_color="#2563EB" if self.selected_service_name == name else ("#E2E8F0", "#374151"),
                text_color=("black", "white") if self.selected_service_name != name else "white",
                corner_radius=10,
                height=34,
                command=lambda n=name: self.select_service(n),
            )
            btn.pack(fill="x", pady=2)

    def select_service(self, service_name: str):
        self.selected_service_name = service_name
        self.refresh_services_panel()

    def add_selected_service(self):
        if not self.selected_service_name:
            messagebox.showwarning("Внимание", "Выберите услугу")
            return
        service = next((s for s in self.service_rows if s["name"] == self.selected_service_name), None)
        if not service:
            return
        try:
            qty = max(1, int(self.qty.get().strip()))
        except ValueError:
            qty = 1
            self.qty.delete(0, "end")
            self.qty.insert(0, "1")
        if self.order_id:
            order_service_id = self.db.add_service_to_order(self.order_id, service["name"], float(service["price"]), qty)
            self.items.append(ServiceItem(order_service_id, service["name"], float(service["price"]), qty))
            self.db.update_order_total(self.order_id)
        else:
            self.items.append(ServiceItem(None, service["name"], float(service["price"]), qty))
        self.refresh_items_panel()

    def refresh_items_panel(self):
        for child in self.items_panel.winfo_children():
            child.destroy()
        if not self.items:
            ctk.CTkLabel(self.items_panel, text="Список услуг пуст", text_color=PALETTE["muted"]).pack(pady=22)
        else:
            header = ctk.CTkFrame(self.items_panel, fg_color=PALETTE["card_alt"], corner_radius=8)
            header.pack(fill="x", pady=(0, 4))
            ctk.CTkLabel(header, text="Услуга", width=290, anchor="w", font=("Arial", 13, "bold")).pack(side="left", padx=8, pady=8)
            ctk.CTkLabel(header, text="Цена", width=90, anchor="e", font=("Arial", 13, "bold")).pack(side="left", padx=6)
            ctk.CTkLabel(header, text="Кол-во", width=80, anchor="center", font=("Arial", 13, "bold")).pack(side="left", padx=6)
            ctk.CTkLabel(header, text="Сумма", width=90, anchor="e", font=("Arial", 13, "bold")).pack(side="left", padx=6)
            for idx, item in enumerate(self.items):
                row = ctk.CTkFrame(self.items_panel, fg_color=PALETTE["card_alt"], corner_radius=8)
                row.pack(fill="x", pady=2)
                ctk.CTkLabel(row, text=item.name, width=290, anchor="w").pack(side="left", padx=8, pady=7)
                ctk.CTkLabel(row, text=f"{item.price:.2f}", width=90, anchor="e").pack(side="left", padx=6)
                qty_entry = ctk.CTkEntry(row, width=80, corner_radius=8)
                qty_entry.insert(0, str(item.quantity))
                qty_entry.pack(side="left", padx=6)
                ctk.CTkLabel(row, text=f"{item.price * item.quantity:.2f}", width=90, anchor="e").pack(side="left", padx=6)
                styled_button(row, "Обнов.", lambda i=idx, e=qty_entry: self.update_qty(i, e), "primary", 76).pack(side="left", padx=3)
                styled_button(row, "Удал.", lambda i=idx: self.remove_item(i), "danger", 68).pack(side="left", padx=3)
        total = sum(item.price * item.quantity for item in self.items)
        self.total_label.configure(text=f"ИТОГО: {total:.2f}")

    def update_qty(self, idx: int, entry: ctk.CTkEntry):
        try:
            value = max(1, int(entry.get().strip()))
        except ValueError:
            value = 1
            entry.delete(0, "end")
            entry.insert(0, "1")
        self.items[idx].quantity = value
        if self.items[idx].service_id:
            self.db.update_order_service_quantity(int(self.items[idx].service_id), value)
            self.db.update_order_total(int(self.order_id))
        self.refresh_items_panel()

    def remove_item(self, idx: int):
        item = self.items[idx]
        if item.service_id:
            self.db.delete_order_service(int(item.service_id))
            self.db.update_order_total(int(self.order_id))
        del self.items[idx]
        self.refresh_items_panel()

    def save_order(self):
        if not self.order_id:
            if not self.client_id:
                selector = ClientSelector(self, self.db)
                selector.grab_set()
                self.wait_window(selector)
                self.client_id = selector.selected_client_id
                if not self.client_id:
                    return
            self.order_id, number = self.db.create_order(self.client_id)
            for item in self.items:
                self.db.add_service_to_order(self.order_id, item.name, item.price, item.quantity)
            total = self.db.update_order_total(self.order_id)
            self.db.update_client_stats(self.client_id, total, 1)
            order = self.db.get_order_by_id(self.order_id)
            self.title_label.configure(text=f"Заказ {number}")
            self.client_label.configure(text=f"Клиент: {order['client_name']} | {order['phone']}")
            self.items = [
                ServiceItem(int(r["id"]), r["service_name"], float(r["price"]), int(r["quantity"]))
                for r in self.db.get_order_services(self.order_id)
            ]
            messagebox.showinfo("Успех", f"Создан заказ {number}")
        else:
            total = self.db.update_order_total(self.order_id)
            order = self.db.get_order_by_id(self.order_id)
            if order and order["client_id"]:
                self.db.cursor.execute(
                    "UPDATE clients SET total_spent = (SELECT COALESCE(SUM(total_sum),0) FROM orders WHERE client_id = ?) WHERE id = ?",
                    (order["client_id"], order["client_id"]),
                )
                self.db.conn.commit()
            messagebox.showinfo("Успех", f"Изменения сохранены, сумма: {total:.2f}")
        self.refresh_items_panel()
        if hasattr(self.master, "refresh_orders"):
            self.master.refresh_orders()

    def print_order(self):
        if not self.order_id:
            messagebox.showwarning("Внимание", "Сначала сохраните заказ")
            return
        order = self.db.get_order_by_id(self.order_id)
        if not order:
            return
        services = self.db.get_order_services(self.order_id)
        default_name = f"zakaz_{order['order_number']}.pdf"
        pdf_path = filedialog.asksaveasfilename(
            title="Сохранить заказ-наряд в PDF",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            initialfile=default_name,
        )
        if not pdf_path:
            return

        try:
            self._save_order_as_pdf(pdf_path, order, services)
        except RuntimeError as err:
            messagebox.showerror("Ошибка", str(err))
            return
        except Exception as err:
            messagebox.showerror("Ошибка", f"Не удалось сохранить PDF:\n{err}")
            return

        if messagebox.askyesno("Печать", "Отправить в печать на принтере?"):
            try:
                self._print_pdf_file(pdf_path)
                messagebox.showinfo("Печать", f"PDF отправлен на печать:\n{pdf_path}")
            except Exception as err:
                messagebox.showerror("Ошибка печати", f"Не удалось отправить PDF на принтер:\n{err}")
        else:
            messagebox.showinfo("Готово", f"PDF сохранен:\n{pdf_path}")

    def _save_order_as_pdf(self, pdf_path: str, order, services):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.pdfgen import canvas
        except ImportError:
            raise RuntimeError("Для сохранения в PDF установите библиотеку reportlab: pip install reportlab")

        font_name = "Helvetica"
        font_candidates = [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/dejavu/DejaVuSans.ttf",
            "C:/Windows/Fonts/arial.ttf",
            "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        ]
        for candidate in font_candidates:
            if os.path.exists(candidate):
                pdfmetrics.registerFont(TTFont("AppFont", candidate))
                font_name = "AppFont"
                break

        c = canvas.Canvas(pdf_path, pagesize=A4)
        width, height = A4
        y = height - 42

        c.setFont(font_name, 16)
        c.drawString(40, y, f"Заказ-наряд {order['order_number']}")
        y -= 24
        c.setFont(font_name, 11)
        c.drawString(40, y, f"Клиент: {order['client_name']} | {format_rf_phone(order['phone'])}")
        y -= 16
        c.drawString(40, y, f"Дата: {order['created_date']}")
        y -= 24

        c.setFont(font_name, 10)
        c.drawString(40, y, "Услуга")
        c.drawString(320, y, "Цена")
        c.drawString(400, y, "Кол-во")
        c.drawString(470, y, "Сумма")
        y -= 10
        c.line(40, y, width - 40, y)
        y -= 14

        for service in services:
            line_total = float(service["price"]) * int(service["quantity"])
            if y < 70:
                c.showPage()
                c.setFont(font_name, 10)
                y = height - 50
            c.drawString(40, y, str(service["service_name"])[:48])
            c.drawRightString(380, y, f"{float(service['price']):.2f}")
            c.drawRightString(440, y, str(int(service["quantity"])))
            c.drawRightString(555, y, f"{line_total:.2f}")
            y -= 14

        y -= 10
        c.line(40, y, width - 40, y)
        y -= 22
        c.setFont(font_name, 12)
        c.drawRightString(width - 40, y, f"ИТОГО: {float(order['total_sum']):.2f}")
        c.save()

    def _print_pdf_file(self, pdf_path: str):
        system_name = platform.system()
        if system_name == "Windows":
            os.startfile(pdf_path, "print")
            return
        subprocess.run(["lp", pdf_path], check=True)


class PeriodManager(ctk.CTkToplevel):
    def __init__(self, parent, db: Database, on_change):
        super().__init__(parent)
        self.db = db
        self.on_change = on_change
        self.title("Периоды цен")
        self.geometry("940x620")
        self.minsize(820, 520)
        self.configure(fg_color=PALETTE["bg"])
        self.setup_ui()
        self.load_periods()

    def setup_ui(self):
        top = card(self)
        top.pack(fill="x", padx=14, pady=(14, 8))
        ctk.CTkLabel(top, text="Управление периодами цен", font=("Arial", 20, "bold")).pack(side="left", padx=12, pady=10)
        actions = ctk.CTkFrame(top, fg_color="transparent")
        actions.pack(side="right", padx=8)
        styled_button(actions, "Новый период", self.create_period, "success", 130).pack(side="left", padx=4, pady=10)
        styled_button(actions, "Импорт Excel", self.import_from_excel, "secondary", 130).pack(side="left", padx=4, pady=10)
        styled_button(actions, "Экспорт", self.export_active_to_excel, "primary", 120).pack(side="left", padx=4, pady=10)
        styled_button(actions, "Закрыть", self.destroy, "ghost", 120).pack(side="left", padx=4, pady=10)

        wrap = card(self)
        wrap.pack(fill="both", expand=True, padx=14, pady=(0, 14))
        self.list_frame = ctk.CTkScrollableFrame(wrap, height=470, fg_color="transparent")
        self.list_frame.pack(fill="both", expand=True, padx=10, pady=10)

    def load_periods(self):
        for child in self.list_frame.winfo_children():
            child.destroy()
        header = ctk.CTkFrame(self.list_frame, fg_color=PALETTE["card_alt"], corner_radius=8)
        header.pack(fill="x", pady=(0, 4))
        ctk.CTkLabel(header, text="ID", width=60, anchor="center", font=("Arial", 13, "bold")).pack(side="left", padx=6, pady=8)
        ctk.CTkLabel(header, text="Название", width=340, anchor="w", font=("Arial", 13, "bold")).pack(side="left", padx=6)
        ctk.CTkLabel(header, text="Дата начала", width=180, anchor="w", font=("Arial", 13, "bold")).pack(side="left", padx=6)
        ctk.CTkLabel(header, text="Статус", width=100, anchor="center", font=("Arial", 13, "bold")).pack(side="left", padx=6)
        ctk.CTkLabel(header, text="Действия", width=220, anchor="center", font=("Arial", 13, "bold")).pack(side="left", padx=6)

        for row in self.db.get_all_periods():
            is_active = bool(row["is_active"])
            frame = ctk.CTkFrame(
                self.list_frame,
                fg_color=("#ECFDF5", "#10261A") if is_active else PALETTE["card_alt"],
                corner_radius=8,
            )
            frame.pack(fill="x", pady=2)
            ctk.CTkLabel(frame, text=str(row["id"]), width=60, anchor="center").pack(side="left", padx=6, pady=7)
            ctk.CTkLabel(frame, text=row["name"], width=340, anchor="w").pack(side="left", padx=6)
            ctk.CTkLabel(frame, text=row["start_date"], width=180, anchor="w").pack(side="left", padx=6)
            ctk.CTkLabel(frame, text="Активен" if is_active else "Завершен", width=100, anchor="center").pack(side="left", padx=6)
            styled_button(frame, "Цены", lambda pid=row["id"]: self.show_prices(pid), "primary", 80).pack(side="left", padx=3)
            if not is_active:
                styled_button(frame, "Активировать", lambda pid=row["id"]: self.activate(pid), "warning", 120).pack(side="left", padx=3)

    def activate(self, period_id: int):
        self.db.activate_period(period_id)
        self.load_periods()
        self.on_change()

    def show_prices(self, period_id: int):
        win = ctk.CTkToplevel(self)
        win.title(f"Цены периода {period_id}")
        win.geometry("620x540")
        win.configure(fg_color=PALETTE["bg"])
        wrap = card(win)
        wrap.pack(fill="both", expand=True, padx=12, pady=12)
        frame = ctk.CTkScrollableFrame(wrap, height=440, fg_color="transparent")
        frame.pack(fill="both", expand=True, padx=10, pady=10)
        for idx, (name, price) in enumerate(self.db.get_period_prices(period_id), 1):
            row = ctk.CTkFrame(frame, fg_color=PALETTE["card_alt"], corner_radius=8)
            row.pack(fill="x", pady=2)
            ctk.CTkLabel(row, text=str(idx), width=50).pack(side="left", padx=4, pady=6)
            ctk.CTkLabel(row, text=name, width=390, anchor="w").pack(side="left", padx=4)
            ctk.CTkLabel(row, text=f"{price:.2f}", width=120, anchor="e").pack(side="left", padx=4)

    def _collect_current_prices(self):
        return [(row["name"], float(row["price"])) for row in self.db.get_active_services()]

    def create_period(self):
        dialog = ctk.CTkInputDialog(text="Название нового периода", title="Новый период")
        name = dialog.get_input()
        if not name:
            return
        prices = self._collect_current_prices()
        if not prices:
            messagebox.showwarning("Ошибка", "Нет услуг для создания периода")
            return
        self.db.create_period_from_prices(name.strip(), prices)
        self.load_periods()
        self.on_change()

    def import_from_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not path:
            return
        wb = load_workbook(path)
        ws = wb.active
        prices: list[tuple[str, float]] = []
        for row in ws.iter_rows(min_row=2):
            if len(row) < 3:
                continue
            name = row[1].value
            value = row[2].value
            if not name or value is None:
                continue
            try:
                prices.append((str(name).strip(), float(str(value).replace(",", "."))))
            except ValueError:
                continue
        if not prices:
            messagebox.showwarning("Ошибка", "Не удалось прочитать цены из файла")
            return
        period_name = f"Период с {datetime.datetime.now().strftime('%d.%m.%Y')}"
        self.db.create_period_from_prices(period_name, prices)
        self.load_periods()
        self.on_change()

    def export_active_to_excel(self):
        prices = self.db.get_period_prices()
        if not prices:
            messagebox.showwarning("Ошибка", "Нет активного периода")
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Цены"
        ws.merge_cells("A1:C1")
        ws["A1"] = "Прайс-лист"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A2"] = f"Дата: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}"
        for col, text in enumerate(["№", "Услуга", "Цена"], 1):
            cell = ws.cell(row=4, column=col, value=text)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for idx, (name, price) in enumerate(prices, 1):
            ws.cell(row=4 + idx, column=1, value=idx)
            ws.cell(row=4 + idx, column=2, value=name)
            ws.cell(row=4 + idx, column=3, value=price)
        ws.column_dimensions[get_column_letter(1)].width = 8
        ws.column_dimensions[get_column_letter(2)].width = 52
        ws.column_dimensions[get_column_letter(3)].width = 18
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"prices_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        )
        if path:
            wb.save(path)
            messagebox.showinfo("Успех", "Файл сохранен")


class StatisticsWindow(ctk.CTkToplevel):
    def __init__(self, parent, db: Database):
        super().__init__(parent)
        self.db = db
        self.period = "week"
        self.title("Статистика")
        self.geometry("1120x720")
        self.minsize(940, 620)
        self.configure(fg_color=PALETTE["bg"])
        self.setup_ui()
        self.load_statistics("week")

    def setup_ui(self):
        top = card(self)
        top.pack(fill="x", padx=14, pady=(14, 8))
        ctk.CTkLabel(top, text="Статистика продаж", font=("Arial", 20, "bold")).pack(side="left", padx=12, pady=10)
        for text, value in [("Неделя", "week"), ("Месяц", "month"), ("Год", "year")]:
            styled_button(top, text, lambda v=value: self.load_statistics(v), "secondary", 110).pack(side="left", padx=4, pady=10)
        self.total_label = ctk.CTkLabel(top, text="Выручка: 0.00", font=("Arial", 14, "bold"), text_color=PALETTE["success"])
        self.total_label.pack(side="right", padx=8)
        self.orders_label = ctk.CTkLabel(top, text="Заказов: 0", font=("Arial", 14, "bold"))
        self.orders_label.pack(side="right", padx=8)
        self.avg_label = ctk.CTkLabel(top, text="Средний чек: 0.00", font=("Arial", 14, "bold"))
        self.avg_label.pack(side="right", padx=8)

        wrap = card(self)
        wrap.pack(fill="both", expand=True, padx=14, pady=(0, 14))
        self.table = ctk.CTkScrollableFrame(wrap, height=560, fg_color="transparent")
        self.table.pack(fill="both", expand=True, padx=10, pady=10)

    def load_statistics(self, period: str):
        self.period = period
        stats = self.db.get_statistics(period)
        self.total_label.configure(text=f"Выручка: {stats['total_sum']:.2f}")
        self.orders_label.configure(text=f"Заказов: {stats['total_orders']}")
        self.avg_label.configure(text=f"Средний чек: {stats['avg_check']:.2f}")
        for child in self.table.winfo_children():
            child.destroy()
        header = ctk.CTkFrame(self.table, fg_color=PALETTE["card_alt"], corner_radius=8)
        header.pack(fill="x", pady=(0, 4))
        ctk.CTkLabel(header, text="Номер", width=140, anchor="w", font=("Arial", 13, "bold")).pack(side="left", padx=8, pady=8)
        ctk.CTkLabel(header, text="Клиент", width=200, anchor="w", font=("Arial", 13, "bold")).pack(side="left", padx=8)
        ctk.CTkLabel(header, text="Телефон", width=170, anchor="w", font=("Arial", 13, "bold")).pack(side="left", padx=8)
        ctk.CTkLabel(header, text="Дата", width=170, anchor="w", font=("Arial", 13, "bold")).pack(side="left", padx=8)
        ctk.CTkLabel(header, text="Сумма", width=120, anchor="e", font=("Arial", 13, "bold")).pack(side="left", padx=8)
        for row in stats["orders"]:
            frame = ctk.CTkFrame(self.table, fg_color=PALETTE["card_alt"], corner_radius=8)
            frame.pack(fill="x", pady=2)
            ctk.CTkLabel(frame, text=row["order_number"], width=140, anchor="w").pack(side="left", padx=8, pady=6)
            ctk.CTkLabel(frame, text=row["client_name"], width=200, anchor="w").pack(side="left", padx=8)
            ctk.CTkLabel(frame, text=row["phone"], width=170, anchor="w").pack(side="left", padx=8)
            ctk.CTkLabel(frame, text=row["created_date"], width=170, anchor="w").pack(side="left", padx=8)
            ctk.CTkLabel(frame, text=f"{row['total_sum']:.2f}", width=120, anchor="e").pack(side="left", padx=8)


class MainApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.db = Database()
        self.orders = []
        self.selected_order_id: int | None = None
        self.title("ИТ-мастерская — Управление заказ-нарядами")
        self.geometry("1320x820")
        self.minsize(1080, 660)
        self.configure(fg_color=PALETTE["bg"])
        self.setup_ui()
        self.refresh_orders()

    def setup_ui(self):
        header = card(self)
        header.pack(fill="x", padx=14, pady=(14, 8))
        title_row = ctk.CTkFrame(header, fg_color="transparent")
        title_row.pack(fill="x", padx=12, pady=(10, 4))
        ctk.CTkLabel(title_row, text="Управление заказ-нарядами", font=("Arial", 24, "bold")).pack(side="left")

        tools_row_1 = ctk.CTkFrame(header, fg_color="transparent")
        tools_row_1.pack(fill="x", padx=12, pady=(0, 4))
        styled_button(tools_row_1, "Периоды цен", self.open_period_manager, "neutral", 128).pack(side="left", padx=4, pady=2)
        styled_button(tools_row_1, "Статистика", self.open_statistics, "neutral", 112).pack(side="left", padx=4, pady=2)
        styled_button(tools_row_1, "Услуги", self.open_service_manager, "neutral", 104).pack(side="left", padx=4, pady=2)
        styled_button(tools_row_1, "Добавить услугу", self.add_service_quick, "neutral", 142).pack(side="left", padx=4, pady=2)
        styled_button(tools_row_1, "Новый клиент", self.create_client_from_main, "neutral", 132).pack(side="left", padx=4, pady=2)

        tools_row_2 = ctk.CTkFrame(header, fg_color="transparent")
        tools_row_2.pack(fill="x", padx=12, pady=(0, 10))
        styled_button(tools_row_2, "Новый заказ", self.create_order, "success", 130).pack(side="left", padx=4, pady=2)
        styled_button(tools_row_2, "Открыть", self.open_selected_order, "neutral", 102).pack(side="left", padx=4, pady=2)
        styled_button(tools_row_2, "Удалить", self.delete_selected_order, "neutral", 102).pack(side="left", padx=4, pady=2)

        body = ctk.CTkFrame(self, fg_color="transparent")
        body.pack(fill="both", expand=True, padx=14, pady=(0, 14))
        left = card(body)
        left.pack(side="left", fill="both", expand=True, padx=(0, 6))
        right = card(body, width=380)
        right.pack(side="left", fill="y", padx=(6, 0))

        search_frame = ctk.CTkFrame(left, fg_color="transparent")
        search_frame.pack(fill="x", padx=10, pady=(10, 6))
        self.search_orders = ctk.CTkEntry(search_frame, placeholder_text="Поиск по номеру, клиенту, телефону", corner_radius=10)
        self.search_orders.pack(side="left", fill="x", expand=True, padx=(0, 6))
        self.search_orders.bind("<KeyRelease>", lambda _: self.display_orders())
        styled_button(search_frame, "Обновить", self.refresh_orders, "ghost", 110).pack(side="left")

        self.orders_panel = ctk.CTkScrollableFrame(left, height=650, fg_color="transparent")
        self.orders_panel.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        ctk.CTkLabel(right, text="Клиенты", font=("Arial", 18, "bold")).pack(anchor="w", padx=12, pady=(12, 6))
        self.clients_panel = ctk.CTkScrollableFrame(right, height=520, fg_color="transparent")
        self.clients_panel.pack(fill="both", expand=True, padx=10, pady=(0, 8))
        self.stats_label = ctk.CTkLabel(right, text="Заказов: 0\nАктивных: 0\nКлиентов: 0", justify="left", font=("Arial", 14))
        self.stats_label.pack(anchor="w", padx=12, pady=(0, 12))

    def refresh_orders(self):
        self.orders = self.db.get_all_orders()
        self.display_orders()
        self.display_clients()
        self.update_stats()

    def display_orders(self):
        for child in self.orders_panel.winfo_children():
            child.destroy()
        query = self.search_orders.get().strip().lower()
        rows = self.orders
        if query:
            rows = [
                row
                for row in rows
                if query in row["order_number"].lower() or query in row["client_name"].lower() or query in row["phone"].lower()
            ]
        header = ctk.CTkFrame(self.orders_panel, fg_color=PALETTE["card_alt"], corner_radius=8)
        header.pack(fill="x", pady=(0, 4))
        ctk.CTkLabel(header, text="№", width=120, anchor="w", font=("Arial", 13, "bold")).pack(side="left", padx=6, pady=8)
        ctk.CTkLabel(header, text="Клиент", width=210, anchor="w", font=("Arial", 13, "bold")).pack(side="left", padx=6)
        ctk.CTkLabel(header, text="Телефон", width=160, anchor="w", font=("Arial", 13, "bold")).pack(side="left", padx=6)
        ctk.CTkLabel(header, text="Дата", width=160, anchor="w", font=("Arial", 13, "bold")).pack(side="left", padx=6)
        ctk.CTkLabel(header, text="Сумма", width=110, anchor="e", font=("Arial", 13, "bold")).pack(side="left", padx=6)
        ctk.CTkLabel(header, text="Статус", width=90, anchor="center", font=("Arial", 13, "bold")).pack(side="left", padx=6)

        for row in rows:
            frame = ctk.CTkFrame(
                self.orders_panel,
                fg_color=("#DBEAFE", "#1E3A5F") if self.selected_order_id == row["id"] else PALETTE["card_alt"],
                corner_radius=8,
            )
            frame.pack(fill="x", pady=2)
            ctk.CTkLabel(frame, text=row["order_number"], width=120, anchor="w").pack(side="left", padx=6, pady=6)
            ctk.CTkLabel(frame, text=row["client_name"], width=210, anchor="w").pack(side="left", padx=6)
            ctk.CTkLabel(frame, text=row["phone"], width=160, anchor="w").pack(side="left", padx=6)
            ctk.CTkLabel(frame, text=row["created_date"], width=160, anchor="w").pack(side="left", padx=6)
            ctk.CTkLabel(frame, text=f"{row['total_sum']:.2f}", width=110, anchor="e").pack(side="left", padx=6)
            ctk.CTkLabel(frame, text=row["status"], width=90, anchor="center").pack(side="left", padx=6)
            styled_button(frame, "Открыть", lambda oid=row["id"]: self.open_order(oid), "primary", 90).pack(side="right", padx=6)
            frame.bind("<Button-1>", lambda _, oid=row["id"]: self.select_order(oid))

    def select_order(self, order_id: int):
        self.selected_order_id = order_id
        self.display_orders()

    def display_clients(self):
        for child in self.clients_panel.winfo_children():
            child.destroy()
        for row in self.db.get_all_clients():
            frame = ctk.CTkFrame(self.clients_panel, fg_color=PALETTE["card_alt"], corner_radius=8)
            frame.pack(fill="x", pady=2)
            ctk.CTkLabel(frame, text=row["name"], width=115, anchor="w").pack(side="left", padx=5, pady=6)
            ctk.CTkLabel(frame, text=format_rf_phone(row["phone"]), width=120, anchor="w").pack(side="left", padx=5)
            ctk.CTkLabel(frame, text=(row["client_comment"] or "")[:16], width=90, anchor="w").pack(side="left", padx=5)
            ctk.CTkLabel(frame, text=str(row["total_orders"]), width=50, anchor="center").pack(side="left", padx=5)
            ctk.CTkLabel(frame, text=f"{row['total_spent']:.2f}", width=90, anchor="e").pack(side="left", padx=5)

    def update_stats(self):
        total = len(self.orders)
        active = len([o for o in self.orders if o["status"] == "active"])
        clients = len(self.db.get_all_clients())
        self.stats_label.configure(text=f"Заказов: {total}\nАктивных: {active}\nКлиентов: {clients}")

    def create_order(self):
        selector = ClientSelector(self, self.db)
        selector.grab_set()
        self.wait_window(selector)
        if selector.selected_client_id:
            editor = OrderEditor(self, self.db, client_id=selector.selected_client_id)
            editor.grab_set()

    def create_client_from_main(self):
        selector = ClientSelector(self, self.db)
        selector.create_client()
        selector.grab_set()
        self.wait_window(selector)
        self.refresh_orders()

    def add_service_quick(self):
        dialog = ctk.CTkToplevel(self)
        dialog.title("Добавить услугу")
        dialog.geometry("520x340")
        dialog.resizable(False, False)
        dialog.configure(fg_color=PALETTE["bg"])
        dialog.grab_set()

        content = card(dialog)
        content.pack(fill="both", expand=True, padx=12, pady=12)
        ctk.CTkLabel(content, text="Новая услуга", font=("Arial", 18, "bold")).pack(anchor="w", padx=12, pady=(12, 6))

        ctk.CTkLabel(content, text="Название услуги").pack(anchor="w", padx=12)
        name_entry = ctk.CTkEntry(content, width=460, corner_radius=10)
        name_entry.pack(padx=12, pady=(2, 8))

        ctk.CTkLabel(content, text="Цена").pack(anchor="w", padx=12)
        price_entry = ctk.CTkEntry(content, width=460, corner_radius=10)
        price_entry.pack(padx=12, pady=(2, 8))

        ctk.CTkLabel(content, text="Категория").pack(anchor="w", padx=12)
        categories = self.db.get_categories()
        category_var = ctk.StringVar(value=categories[0] if categories else "Основные")
        category_menu = ctk.CTkOptionMenu(content, values=categories if categories else ["Основные"], variable=category_var, width=330)
        category_menu.pack(anchor="w", padx=12, pady=(2, 8))

        def add_new_category():
            category_dialog = ctk.CTkInputDialog(text="Название новой категории", title="Новая категория")
            new_cat = (category_dialog.get_input() or "").strip()
            if not new_cat:
                return
            current = list(category_menu.cget("values"))
            if new_cat not in current:
                current.append(new_cat)
                category_menu.configure(values=current)
            category_var.set(new_cat)

        styled_button(content, "Добавить категорию", add_new_category, "secondary", 170).pack(anchor="w", padx=12, pady=(0, 10))

        def save():
            name = name_entry.get().strip()
            category = category_var.get().strip()
            try:
                price = float(price_entry.get().strip().replace(",", "."))
            except ValueError:
                messagebox.showwarning("Ошибка", "Введите корректную цену")
                return
            if not name or not category or price < 0:
                messagebox.showwarning("Ошибка", "Заполните поля корректно")
                return
            if self.db.get_service_by_name(name):
                messagebox.showwarning("Ошибка", "Услуга с таким названием уже существует")
                return
            if not self.db.add_service(name, price, category):
                messagebox.showwarning("Ошибка", "Не удалось добавить услугу")
                return
            dialog.destroy()
            self.refresh_orders()
            messagebox.showinfo("Успех", "Услуга добавлена")

        buttons = ctk.CTkFrame(content, fg_color="transparent")
        buttons.pack(fill="x", padx=12, pady=(0, 12))
        styled_button(buttons, "Сохранить", save, "success", 120).pack(side="right", padx=6)
        styled_button(buttons, "Отмена", dialog.destroy, "ghost", 110).pack(side="right", padx=6)

    def open_order(self, order_id: int):
        editor = OrderEditor(self, self.db, order_id=order_id)
        editor.grab_set()

    def open_selected_order(self):
        if not self.selected_order_id:
            messagebox.showwarning("Внимание", "Заказ не выбран")
            return
        self.open_order(self.selected_order_id)

    def delete_selected_order(self):
        if not self.selected_order_id:
            messagebox.showwarning("Внимание", "Заказ не выбран")
            return
        if not messagebox.askyesno("Подтверждение", "Удалить выбранный заказ?"):
            return
        self.db.delete_order(self.selected_order_id)
        self.selected_order_id = None
        self.refresh_orders()

    def open_service_manager(self):
        manager = ServiceManager(self, self.db, self.refresh_orders)
        manager.grab_set()

    def open_statistics(self):
        win = StatisticsWindow(self, self.db)
        win.grab_set()

    def open_period_manager(self):
        win = PeriodManager(self, self.db, self.refresh_orders)
        win.grab_set()

    def on_close(self):
        self.db.close()
        self.destroy()

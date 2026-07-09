import datetime
import sqlite3
import tempfile
import webbrowser
from dataclasses import dataclass
from tkinter import filedialog, messagebox

import customtkinter as ctk
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")


@dataclass
class ServiceItem:
    service_id: int | None
    name: str
    price: float
    quantity: int


class Database:
    def __init__(self, db_name: str = "orders.db"):
        self.conn = sqlite3.connect(db_name)
        self.conn.execute("PRAGMA foreign_keys = ON")
        self.conn.row_factory = sqlite3.Row
        self.cursor = self.conn.cursor()
        self.current_period_id: int | None = None
        self.create_tables()
        self.seed_data()
        self.load_current_period()

    def _now(self) -> str:
        return datetime.datetime.now().strftime("%d.%m.%Y %H:%M")

    def create_tables(self) -> None:
        self.cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS clients (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                phone TEXT NOT NULL UNIQUE,
                created_date TEXT NOT NULL,
                total_orders INTEGER DEFAULT 0,
                total_spent REAL DEFAULT 0
            )
            """
        )
        self.cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS services_catalog (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                price REAL NOT NULL,
                category TEXT NOT NULL DEFAULT 'Основные',
                created_date TEXT NOT NULL,
                is_active INTEGER NOT NULL DEFAULT 1
            )
            """
        )
        self.cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS price_periods (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                start_date TEXT NOT NULL,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_date TEXT NOT NULL
            )
            """
        )
        self.cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS period_prices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                period_id INTEGER NOT NULL,
                service_name TEXT NOT NULL,
                price REAL NOT NULL,
                FOREIGN KEY (period_id) REFERENCES price_periods(id) ON DELETE CASCADE
            )
            """
        )
        self.cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_number TEXT NOT NULL UNIQUE,
                client_id INTEGER,
                created_date TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'active',
                total_sum REAL NOT NULL DEFAULT 0,
                period_id INTEGER,
                FOREIGN KEY (client_id) REFERENCES clients(id),
                FOREIGN KEY (period_id) REFERENCES price_periods(id)
            )
            """
        )
        self.cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS order_services (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id INTEGER NOT NULL,
                service_name TEXT NOT NULL,
                price REAL NOT NULL,
                quantity INTEGER NOT NULL DEFAULT 1,
                FOREIGN KEY (order_id) REFERENCES orders(id) ON DELETE CASCADE
            )
            """
        )
        self.conn.commit()

    def seed_data(self) -> None:
        self.cursor.execute("SELECT COUNT(*) FROM services_catalog")
        if self.cursor.fetchone()[0] == 0:
            now = self._now()
            services = [
                ("Диагностика (вычитается из ремонта)", 500, "Диагностика"),
                ("Сборка ПК под ключ", 3000, "Компьютеры"),
                ("Апгрейд ПК", 1500, "Компьютеры"),
                ("Профилактика ПК (чистка + термопаста)", 2000, "Компьютеры"),
                ("Чистка ноутбука + термопаста", 2500, "Ноутбуки"),
                ("Замена матрицы ноутбука", 3000, "Ноутбуки"),
                ("Установка Windows 10/11 + драйверы", 2500, "Программное обеспечение"),
                ("Удаление вирусов", 2000, "Программное обеспечение"),
                ("Выезд мастера + диагностика", 1500, "Выездные услуги"),
                ("3D-печать (стандартная)", 500, "3D-печать"),
            ]
            self.cursor.executemany(
                "INSERT INTO services_catalog (name, price, category, created_date, is_active) VALUES (?, ?, ?, ?, 1)",
                [(name, price, category, now) for name, price, category in services],
            )
        self.cursor.execute("SELECT COUNT(*) FROM price_periods")
        if self.cursor.fetchone()[0] == 0:
            self.create_period_from_prices(
                f"Период с {datetime.datetime.now().strftime('%d.%m.%Y')}",
                [(s["name"], s["price"]) for s in self.get_active_services()],
            )
        self.conn.commit()

    def load_current_period(self) -> None:
        self.cursor.execute("SELECT id FROM price_periods WHERE is_active = 1 ORDER BY id DESC LIMIT 1")
        row = self.cursor.fetchone()
        self.current_period_id = row["id"] if row else None

    def get_active_services(self) -> list[sqlite3.Row]:
        self.cursor.execute(
            "SELECT id, name, price, category, is_active FROM services_catalog WHERE is_active = 1 ORDER BY category, name"
        )
        return self.cursor.fetchall()

    def get_all_services(self) -> list[sqlite3.Row]:
        self.cursor.execute("SELECT id, name, price, category, is_active FROM services_catalog ORDER BY category, name")
        return self.cursor.fetchall()

    def get_categories(self) -> list[str]:
        self.cursor.execute("SELECT DISTINCT category FROM services_catalog ORDER BY category")
        return [row["category"] for row in self.cursor.fetchall()]

    def get_service_by_name(self, name: str) -> sqlite3.Row | None:
        self.cursor.execute("SELECT id, name, price, category FROM services_catalog WHERE name = ?", (name,))
        return self.cursor.fetchone()

    def add_service(self, name: str, price: float, category: str) -> int | None:
        try:
            self.cursor.execute(
                "INSERT INTO services_catalog (name, price, category, created_date, is_active) VALUES (?, ?, ?, ?, 1)",
                (name, price, category, self._now()),
            )
            self.conn.commit()
            return self.cursor.lastrowid
        except sqlite3.IntegrityError:
            return None

    def update_service(self, service_id: int, name: str, price: float, category: str) -> bool:
        try:
            self.cursor.execute(
                "UPDATE services_catalog SET name = ?, price = ?, category = ? WHERE id = ?",
                (name, price, category, service_id),
            )
            self.conn.commit()
            return self.cursor.rowcount > 0
        except sqlite3.IntegrityError:
            return False

    def set_service_active(self, service_id: int, active: bool) -> None:
        self.cursor.execute("UPDATE services_catalog SET is_active = ? WHERE id = ?", (1 if active else 0, service_id))
        self.conn.commit()

    def delete_service(self, service_id: int) -> None:
        self.cursor.execute("DELETE FROM services_catalog WHERE id = ?", (service_id,))
        self.conn.commit()

    def get_all_clients(self) -> list[sqlite3.Row]:
        self.cursor.execute(
            "SELECT id, name, phone, created_date, total_orders, total_spent FROM clients ORDER BY id DESC"
        )
        return self.cursor.fetchall()

    def search_clients(self, query: str) -> list[sqlite3.Row]:
        like = f"%{query.strip()}%"
        self.cursor.execute(
            "SELECT id, name, phone, created_date, total_orders, total_spent FROM clients WHERE name LIKE ? OR phone LIKE ? ORDER BY id DESC",
            (like, like),
        )
        return self.cursor.fetchall()

    def find_client_by_name_phone(self, name: str, phone: str) -> sqlite3.Row | None:
        self.cursor.execute("SELECT id, name, phone FROM clients WHERE name = ? AND phone = ?", (name, phone))
        return self.cursor.fetchone()

    def create_client(self, name: str, phone: str) -> int | None:
        try:
            self.cursor.execute(
                "INSERT INTO clients (name, phone, created_date, total_orders, total_spent) VALUES (?, ?, ?, 0, 0)",
                (name, phone, self._now()),
            )
            self.conn.commit()
            return self.cursor.lastrowid
        except sqlite3.IntegrityError:
            return None

    def update_client_stats(self, client_id: int, delta_sum: float, delta_orders: int = 1) -> None:
        self.cursor.execute(
            "UPDATE clients SET total_orders = total_orders + ?, total_spent = total_spent + ? WHERE id = ?",
            (delta_orders, delta_sum, client_id),
        )
        self.conn.commit()

    def get_client_orders(self, client_id: int) -> list[sqlite3.Row]:
        self.cursor.execute(
            "SELECT id, order_number, created_date, status, total_sum FROM orders WHERE client_id = ? ORDER BY id DESC",
            (client_id,),
        )
        return self.cursor.fetchall()

    def get_all_orders(self) -> list[sqlite3.Row]:
        self.cursor.execute(
            """
            SELECT o.id, o.order_number, COALESCE(c.name, 'Без клиента') AS client_name, COALESCE(c.phone, '') AS phone,
                   o.created_date, o.status, o.total_sum
            FROM orders o
            LEFT JOIN clients c ON c.id = o.client_id
            ORDER BY o.id DESC
            """
        )
        return self.cursor.fetchall()

    def get_order_by_id(self, order_id: int) -> sqlite3.Row | None:
        self.cursor.execute(
            """
            SELECT o.id, o.order_number, COALESCE(c.name, 'Без клиента') AS client_name, COALESCE(c.phone, '') AS phone,
                   o.created_date, o.status, o.total_sum, o.client_id
            FROM orders o
            LEFT JOIN clients c ON c.id = o.client_id
            WHERE o.id = ?
            """,
            (order_id,),
        )
        return self.cursor.fetchone()

    def next_order_number(self) -> str:
        self.cursor.execute("SELECT order_number FROM orders ORDER BY id DESC LIMIT 1")
        row = self.cursor.fetchone()
        if not row:
            return "ORD-000001"
        try:
            value = int(str(row["order_number"]).split("-")[1]) + 1
            return f"ORD-{value:06d}"
        except Exception:
            return "ORD-000001"

    def create_order(self, client_id: int | None) -> tuple[int, str]:
        number = self.next_order_number()
        self.cursor.execute(
            "INSERT INTO orders (order_number, client_id, created_date, status, total_sum, period_id) VALUES (?, ?, ?, 'active', 0, ?)",
            (number, client_id, self._now(), self.current_period_id),
        )
        self.conn.commit()
        return self.cursor.lastrowid, number

    def get_order_services(self, order_id: int) -> list[sqlite3.Row]:
        self.cursor.execute(
            "SELECT id, service_name, price, quantity FROM order_services WHERE order_id = ? ORDER BY id",
            (order_id,),
        )
        return self.cursor.fetchall()

    def add_service_to_order(self, order_id: int, service_name: str, price: float, quantity: int = 1) -> int:
        self.cursor.execute(
            "INSERT INTO order_services (order_id, service_name, price, quantity) VALUES (?, ?, ?, ?)",
            (order_id, service_name, price, quantity),
        )
        self.conn.commit()
        return self.cursor.lastrowid

    def update_order_service_quantity(self, order_service_id: int, quantity: int) -> None:
        self.cursor.execute("UPDATE order_services SET quantity = ? WHERE id = ?", (quantity, order_service_id))
        self.conn.commit()

    def delete_order_service(self, order_service_id: int) -> None:
        self.cursor.execute("DELETE FROM order_services WHERE id = ?", (order_service_id,))
        self.conn.commit()

    def update_order_total(self, order_id: int) -> float:
        self.cursor.execute("SELECT COALESCE(SUM(price * quantity), 0) AS total FROM order_services WHERE order_id = ?", (order_id,))
        total = float(self.cursor.fetchone()["total"])
        self.cursor.execute("UPDATE orders SET total_sum = ? WHERE id = ?", (total, order_id))
        self.conn.commit()
        return total

    def delete_order(self, order_id: int) -> None:
        self.cursor.execute("SELECT client_id, total_sum FROM orders WHERE id = ?", (order_id,))
        row = self.cursor.fetchone()
        if row and row["client_id"]:
            self.update_client_stats(int(row["client_id"]), -float(row["total_sum"]), -1)
        self.cursor.execute("DELETE FROM orders WHERE id = ?", (order_id,))
        self.conn.commit()

    def set_order_client(self, order_id: int, client_id: int) -> None:
        self.cursor.execute("UPDATE orders SET client_id = ? WHERE id = ?", (client_id, order_id))
        self.conn.commit()

    def get_statistics(self, period: str) -> dict:
        now = datetime.datetime.now()
        if period == "month":
            start = now.replace(day=1, hour=0, minute=0).strftime("%d.%m.%Y")
        elif period == "year":
            start = now.replace(month=1, day=1, hour=0, minute=0).strftime("%d.%m.%Y")
        else:
            start = (now - datetime.timedelta(days=7)).strftime("%d.%m.%Y")
        self.cursor.execute(
            """
            SELECT o.id, o.order_number, o.created_date, o.total_sum, COALESCE(c.name, 'Без клиента') AS client_name, COALESCE(c.phone, '') AS phone
            FROM orders o
            LEFT JOIN clients c ON c.id = o.client_id
            WHERE o.created_date >= ?
            ORDER BY o.id DESC
            """,
            (start,),
        )
        orders = self.cursor.fetchall()
        total_sum = sum(float(o["total_sum"]) for o in orders)
        total_orders = len(orders)
        avg_check = total_sum / total_orders if total_orders else 0
        return {"total_sum": total_sum, "total_orders": total_orders, "avg_check": avg_check, "orders": orders}

    def get_top_clients_by_orders(self, min_orders: int = 2) -> list[sqlite3.Row]:
        self.cursor.execute(
            "SELECT id, name, phone, total_orders, total_spent FROM clients WHERE total_orders >= ? ORDER BY total_orders DESC",
            (min_orders,),
        )
        return self.cursor.fetchall()

    def get_top_clients_by_spent(self, limit: int = 10) -> list[sqlite3.Row]:
        self.cursor.execute(
            "SELECT id, name, phone, total_orders, total_spent FROM clients WHERE total_orders > 0 ORDER BY total_spent DESC LIMIT ?",
            (limit,),
        )
        return self.cursor.fetchall()

    def get_all_periods(self) -> list[sqlite3.Row]:
        self.cursor.execute("SELECT id, name, start_date, is_active FROM price_periods ORDER BY id DESC")
        return self.cursor.fetchall()

    def get_period_prices(self, period_id: int | None = None) -> list[tuple[str, float]]:
        pid = period_id if period_id is not None else self.current_period_id
        if pid is None:
            return []
        self.cursor.execute("SELECT service_name, price FROM period_prices WHERE period_id = ? ORDER BY service_name", (pid,))
        return [(row["service_name"], float(row["price"])) for row in self.cursor.fetchall()]

    def create_period_from_prices(self, period_name: str, prices: list[tuple[str, float]]) -> int:
        now = self._now()
        self.cursor.execute("UPDATE price_periods SET is_active = 0 WHERE is_active = 1")
        self.cursor.execute(
            "INSERT INTO price_periods (name, start_date, is_active, created_date) VALUES (?, ?, 1, ?)",
            (period_name, now, now),
        )
        period_id = self.cursor.lastrowid
        self.cursor.execute("DELETE FROM period_prices WHERE period_id = ?", (period_id,))
        for service_name, price in prices:
            self.cursor.execute(
                "INSERT INTO period_prices (period_id, service_name, price) VALUES (?, ?, ?)",
                (period_id, service_name, float(price)),
            )
        self.cursor.execute("DELETE FROM services_catalog")
        for service_name, price in prices:
            self.cursor.execute(
                "INSERT INTO services_catalog (name, price, category, created_date, is_active) VALUES (?, ?, 'Основные', ?, 1)",
                (service_name, float(price), now),
            )
        self.conn.commit()
        self.current_period_id = period_id
        return period_id

    def activate_period(self, period_id: int) -> None:
        self.cursor.execute("UPDATE price_periods SET is_active = 0 WHERE is_active = 1")
        self.cursor.execute("UPDATE price_periods SET is_active = 1 WHERE id = ?", (period_id,))
        self.conn.commit()
        self.current_period_id = period_id

    def close(self) -> None:
        self.conn.close()


class ServiceManager(ctk.CTkToplevel):
    def __init__(self, parent, db: Database, on_change):
        super().__init__(parent)
        self.db = db
        self.on_change = on_change
        self.title("Управление услугами")
        self.geometry("900x560")
        self.setup_ui()
        self.load_services()

    def setup_ui(self) -> None:
        top = ctk.CTkFrame(self)
        top.pack(fill="x", padx=10, pady=10)
        ctk.CTkButton(top, text="Добавить", command=self.add_service_dialog, fg_color="green", width=120).pack(side="left", padx=5)
        ctk.CTkButton(top, text="Обновить", command=self.load_services, width=120).pack(side="left", padx=5)
        ctk.CTkButton(top, text="Закрыть", command=self.destroy, fg_color="gray", width=120).pack(side="right", padx=5)
        self.table = ctk.CTkScrollableFrame(self, height=440)
        self.table.pack(fill="both", expand=True, padx=10, pady=5)

    def load_services(self) -> None:
        for child in self.table.winfo_children():
            child.destroy()
        header = ctk.CTkFrame(self.table, fg_color="transparent")
        header.pack(fill="x", pady=(0, 2))
        ctk.CTkLabel(header, text="Название", width=340, anchor="w").pack(side="left", padx=5)
        ctk.CTkLabel(header, text="Цена", width=120, anchor="e").pack(side="left", padx=5)
        ctk.CTkLabel(header, text="Категория", width=160, anchor="w").pack(side="left", padx=5)
        ctk.CTkLabel(header, text="Статус", width=90, anchor="center").pack(side="left", padx=5)
        ctk.CTkLabel(header, text="Действия", width=240, anchor="center").pack(side="left", padx=5)
        for row in self.db.get_all_services():
            frame = ctk.CTkFrame(self.table, fg_color="transparent")
            frame.pack(fill="x", pady=1)
            active = bool(row["is_active"])
            ctk.CTkLabel(frame, text=row["name"], width=340, anchor="w").pack(side="left", padx=5)
            ctk.CTkLabel(frame, text=f"{row['price']:.2f}", width=120, anchor="e").pack(side="left", padx=5)
            ctk.CTkLabel(frame, text=row["category"], width=160, anchor="w").pack(side="left", padx=5)
            ctk.CTkLabel(frame, text="Активна" if active else "Отключена", width=90, anchor="center").pack(side="left", padx=5)
            ctk.CTkButton(frame, text="Редакт.", width=70, command=lambda sid=row["id"]: self.edit_service_dialog(sid)).pack(side="left", padx=3)
            ctk.CTkButton(
                frame,
                text="Откл." if active else "Вкл.",
                width=70,
                fg_color="orange",
                command=lambda sid=row["id"], st=active: self.toggle_service(sid, st),
            ).pack(side="left", padx=3)
            ctk.CTkButton(frame, text="Удал.", width=70, fg_color="red", command=lambda sid=row["id"]: self.delete_service(sid)).pack(side="left", padx=3)

    def _service_dialog(self, title: str, service=None) -> tuple[str, float, str] | None:
        dialog = ctk.CTkInputDialog(text=f"{title}\nВведите: название|цена|категория", title=title)
        value = dialog.get_input()
        if not value:
            return None
        parts = [part.strip() for part in value.split("|")]
        if len(parts) < 2:
            messagebox.showwarning("Ошибка", "Формат: название|цена|категория")
            return None
        name = parts[0]
        try:
            price = float(parts[1].replace(",", "."))
        except ValueError:
            messagebox.showwarning("Ошибка", "Некорректная цена")
            return None
        category = parts[2] if len(parts) > 2 and parts[2] else (service["category"] if service else "Основные")
        if not name or price < 0:
            messagebox.showwarning("Ошибка", "Проверьте данные")
            return None
        return name, price, category

    def add_service_dialog(self) -> None:
        data = self._service_dialog("Добавить услугу")
        if not data:
            return
        name, price, category = data
        if not self.db.add_service(name, price, category):
            messagebox.showwarning("Ошибка", "Услуга с таким названием уже есть")
            return
        self.load_services()
        self.on_change()

    def edit_service_dialog(self, service_id: int) -> None:
        service = next((s for s in self.db.get_all_services() if s["id"] == service_id), None)
        if not service:
            return
        data = self._service_dialog("Редактировать услугу", service)
        if not data:
            return
        name, price, category = data
        if not self.db.update_service(service_id, name, price, category):
            messagebox.showwarning("Ошибка", "Не удалось сохранить")
            return
        self.load_services()
        self.on_change()

    def toggle_service(self, service_id: int, current_active: bool) -> None:
        self.db.set_service_active(service_id, not current_active)
        self.load_services()
        self.on_change()

    def delete_service(self, service_id: int) -> None:
        if not messagebox.askyesno("Подтверждение", "Удалить услугу?"):
            return
        self.db.delete_service(service_id)
        self.load_services()
        self.on_change()


class ClientSelector(ctk.CTkToplevel):
    def __init__(self, parent, db: Database):
        super().__init__(parent)
        self.db = db
        self.selected_client_id: int | None = None
        self.title("Выбор клиента")
        self.geometry("640x520")
        self.setup_ui()
        self.load_clients()

    def setup_ui(self) -> None:
        top = ctk.CTkFrame(self)
        top.pack(fill="x", padx=10, pady=10)
        self.search = ctk.CTkEntry(top, placeholder_text="Имя или телефон")
        self.search.pack(side="left", fill="x", expand=True, padx=5)
        self.search.bind("<KeyRelease>", lambda _: self.load_clients())
        ctk.CTkButton(top, text="Новый клиент", command=self.create_client, fg_color="green", width=140).pack(side="left", padx=5)
        ctk.CTkButton(top, text="Выбрать", command=self.confirm, width=120).pack(side="left", padx=5)
        self.list_frame = ctk.CTkScrollableFrame(self, height=410)
        self.list_frame.pack(fill="both", expand=True, padx=10, pady=5)

    def load_clients(self) -> None:
        for child in self.list_frame.winfo_children():
            child.destroy()
        query = self.search.get().strip()
        rows = self.db.search_clients(query) if query else self.db.get_all_clients()
        for row in rows:
            frame = ctk.CTkFrame(self.list_frame, fg_color="transparent")
            frame.pack(fill="x", pady=1)
            ctk.CTkLabel(frame, text=row["name"], width=180, anchor="w").pack(side="left", padx=5)
            ctk.CTkLabel(frame, text=row["phone"], width=150, anchor="w").pack(side="left", padx=5)
            ctk.CTkLabel(frame, text=str(row["total_orders"]), width=80, anchor="center").pack(side="left", padx=5)
            ctk.CTkLabel(frame, text=f"{row['total_spent']:.2f}", width=120, anchor="e").pack(side="left", padx=5)
            ctk.CTkButton(frame, text="Выбрать", width=90, command=lambda cid=row["id"]: self.select(cid)).pack(side="right", padx=5)

    def create_client(self) -> None:
        dialog = ctk.CTkInputDialog(text="Введите: имя|телефон", title="Новый клиент")
        value = dialog.get_input()
        if not value:
            return
        parts = [p.strip() for p in value.split("|")]
        if len(parts) < 2:
            messagebox.showwarning("Ошибка", "Формат: имя|телефон")
            return
        name, phone = parts[0], parts[1]
        existing = self.db.find_client_by_name_phone(name, phone)
        if existing:
            self.select(int(existing["id"]))
            return
        client_id = self.db.create_client(name, phone)
        if client_id is None:
            messagebox.showwarning("Ошибка", "Не удалось создать клиента")
            return
        self.select(client_id)

    def select(self, client_id: int) -> None:
        self.selected_client_id = client_id
        self.confirm()

    def confirm(self) -> None:
        if self.selected_client_id is None:
            messagebox.showwarning("Внимание", "Клиент не выбран")
            return
        self.destroy()


class OrderEditor(ctk.CTkToplevel):
    def __init__(self, parent, db: Database, order_id: int | None = None, client_id: int | None = None):
        super().__init__(parent)
        self.db = db
        self.order_id = order_id
        self.client_id = client_id
        self.items: list[ServiceItem] = []
        self.service_rows = self.db.get_active_services()
        self.selected_service_name: str | None = None
        self.title("Заказ-наряд")
        self.geometry("1120x720")
        self.setup_ui()
        if self.order_id:
            self.load_existing_order()
        self.refresh_services_panel()
        self.refresh_items_panel()

    def setup_ui(self) -> None:
        top = ctk.CTkFrame(self)
        top.pack(fill="x", padx=10, pady=8)
        self.title_label = ctk.CTkLabel(top, text="Новый заказ", font=("Arial", 18, "bold"))
        self.title_label.pack(side="left", padx=10)
        self.client_label = ctk.CTkLabel(top, text="Клиент: не выбран", font=("Arial", 14))
        self.client_label.pack(side="left", padx=15)

        body = ctk.CTkFrame(self)
        body.pack(fill="both", expand=True, padx=10, pady=5)

        left = ctk.CTkFrame(body)
        left.pack(side="left", fill="both", expand=True, padx=(0, 5))
        ctk.CTkLabel(left, text="Каталог услуг", font=("Arial", 14, "bold")).pack(pady=5)
        search_frame = ctk.CTkFrame(left)
        search_frame.pack(fill="x", padx=5, pady=5)
        self.search_service = ctk.CTkEntry(search_frame, placeholder_text="Поиск услуги")
        self.search_service.pack(side="left", fill="x", expand=True, padx=(0, 5))
        self.search_service.bind("<KeyRelease>", lambda _: self.refresh_services_panel())
        ctk.CTkButton(search_frame, text="Услуги", width=90, command=self.open_service_manager).pack(side="left")
        self.services_panel = ctk.CTkScrollableFrame(left, height=420)
        self.services_panel.pack(fill="both", expand=True, padx=5, pady=5)

        add_frame = ctk.CTkFrame(left)
        add_frame.pack(fill="x", padx=5, pady=5)
        ctk.CTkLabel(add_frame, text="Кол-во").pack(side="left", padx=5)
        self.qty = ctk.CTkEntry(add_frame, width=60)
        self.qty.insert(0, "1")
        self.qty.pack(side="left", padx=5)
        ctk.CTkButton(add_frame, text="Добавить", command=self.add_selected_service, fg_color="green").pack(side="left", padx=8)

        right = ctk.CTkFrame(body)
        right.pack(side="left", fill="both", expand=True, padx=(5, 0))
        ctk.CTkLabel(right, text="Услуги в заказе", font=("Arial", 14, "bold")).pack(pady=5)
        self.items_panel = ctk.CTkScrollableFrame(right, height=420)
        self.items_panel.pack(fill="both", expand=True, padx=5, pady=5)
        bottom = ctk.CTkFrame(right)
        bottom.pack(fill="x", padx=5, pady=5)
        self.total_label = ctk.CTkLabel(bottom, text="ИТОГО: 0.00", font=("Arial", 18, "bold"))
        self.total_label.pack(side="left", padx=8)
        ctk.CTkButton(bottom, text="Печать", command=self.print_order, fg_color="orange", width=120).pack(side="right", padx=5)
        ctk.CTkButton(bottom, text="Сохранить", command=self.save_order, fg_color="green", width=120).pack(side="right", padx=5)

    def load_existing_order(self) -> None:
        order = self.db.get_order_by_id(self.order_id)
        if not order:
            return
        self.client_id = order["client_id"]
        self.title_label.configure(text=f"Заказ {order['order_number']}")
        self.client_label.configure(text=f"Клиент: {order['client_name']} | {order['phone']}")
        rows = self.db.get_order_services(self.order_id)
        self.items = [ServiceItem(int(r["id"]), r["service_name"], float(r["price"]), int(r["quantity"])) for r in rows]

    def open_service_manager(self) -> None:
        manager = ServiceManager(self, self.db, self.reload_services)
        manager.grab_set()

    def reload_services(self) -> None:
        self.service_rows = self.db.get_active_services()
        self.refresh_services_panel()

    def refresh_services_panel(self) -> None:
        for child in self.services_panel.winfo_children():
            child.destroy()
        query = self.search_service.get().strip().lower()
        rows = self.service_rows
        if query:
            rows = [s for s in rows if query in s["name"].lower()]
        for row in rows:
            name = row["name"]
            btn = ctk.CTkButton(
                self.services_panel,
                text=f"{name} — {row['price']:.2f}",
                anchor="w",
                fg_color=("#3B8ED0", "#1F6AA5") if self.selected_service_name == name else "transparent",
                hover_color=("#E0E0E0", "#3A3A3A"),
                command=lambda n=name: self.select_service(n),
            )
            btn.pack(fill="x", pady=1)

    def select_service(self, service_name: str) -> None:
        self.selected_service_name = service_name
        self.refresh_services_panel()

    def add_selected_service(self) -> None:
        if not self.selected_service_name:
            messagebox.showwarning("Внимание", "Выберите услугу")
            return
        service = next((s for s in self.service_rows if s["name"] == self.selected_service_name), None)
        if not service:
            return
        try:
            qty = max(1, int(self.qty.get().strip()))
        except ValueError:
            qty = 1
            self.qty.delete(0, "end")
            self.qty.insert(0, "1")
        if self.order_id:
            order_service_id = self.db.add_service_to_order(self.order_id, service["name"], float(service["price"]), qty)
            self.items.append(ServiceItem(order_service_id, service["name"], float(service["price"]), qty))
            self.db.update_order_total(self.order_id)
        else:
            self.items.append(ServiceItem(None, service["name"], float(service["price"]), qty))
        self.refresh_items_panel()

    def refresh_items_panel(self) -> None:
        for child in self.items_panel.winfo_children():
            child.destroy()
        if not self.items:
            ctk.CTkLabel(self.items_panel, text="Список пуст", text_color="gray").pack(pady=20)
        for idx, item in enumerate(self.items):
            row = ctk.CTkFrame(self.items_panel, fg_color="transparent")
            row.pack(fill="x", pady=1)
            ctk.CTkLabel(row, text=item.name, width=280, anchor="w").pack(side="left", padx=5)
            ctk.CTkLabel(row, text=f"{item.price:.2f}", width=90, anchor="e").pack(side="left", padx=5)
            qty_entry = ctk.CTkEntry(row, width=60)
            qty_entry.insert(0, str(item.quantity))
            qty_entry.pack(side="left", padx=5)
            ctk.CTkLabel(row, text=f"{item.price * item.quantity:.2f}", width=90, anchor="e").pack(side="left", padx=5)
            ctk.CTkButton(row, text="Обнов.", width=70, command=lambda i=idx, e=qty_entry: self.update_qty(i, e)).pack(side="left", padx=3)
            ctk.CTkButton(row, text="Удал.", width=70, fg_color="red", command=lambda i=idx: self.remove_item(i)).pack(side="left", padx=3)
        total = sum(item.price * item.quantity for item in self.items)
        self.total_label.configure(text=f"ИТОГО: {total:.2f}")

    def update_qty(self, idx: int, entry: ctk.CTkEntry) -> None:
        try:
            value = max(1, int(entry.get().strip()))
        except ValueError:
            value = 1
            entry.delete(0, "end")
            entry.insert(0, "1")
        self.items[idx].quantity = value
        if self.items[idx].service_id:
            self.db.update_order_service_quantity(int(self.items[idx].service_id), value)
            self.db.update_order_total(int(self.order_id))
        self.refresh_items_panel()

    def remove_item(self, idx: int) -> None:
        item = self.items[idx]
        if item.service_id:
            self.db.delete_order_service(int(item.service_id))
            self.db.update_order_total(int(self.order_id))
        del self.items[idx]
        self.refresh_items_panel()

    def save_order(self) -> None:
        if not self.order_id:
            if not self.client_id:
                selector = ClientSelector(self, self.db)
                selector.grab_set()
                self.wait_window(selector)
                self.client_id = selector.selected_client_id
                if not self.client_id:
                    return
            self.order_id, number = self.db.create_order(self.client_id)
            for item in self.items:
                self.db.add_service_to_order(self.order_id, item.name, item.price, item.quantity)
            total = self.db.update_order_total(self.order_id)
            self.db.update_client_stats(self.client_id, total, 1)
            order = self.db.get_order_by_id(self.order_id)
            self.title_label.configure(text=f"Заказ {number}")
            self.client_label.configure(text=f"Клиент: {order['client_name']} | {order['phone']}")
            self.items = [ServiceItem(int(r["id"]), r["service_name"], float(r["price"]), int(r["quantity"])) for r in self.db.get_order_services(self.order_id)]
            messagebox.showinfo("Успех", f"Создан заказ {number}")
        else:
            total = self.db.update_order_total(self.order_id)
            order = self.db.get_order_by_id(self.order_id)
            if order and order["client_id"]:
                self.db.cursor.execute(
                    "UPDATE clients SET total_spent = (SELECT COALESCE(SUM(total_sum),0) FROM orders WHERE client_id = ?) WHERE id = ?",
                    (order["client_id"], order["client_id"]),
                )
                self.db.conn.commit()
            messagebox.showinfo("Успех", f"Изменения сохранены, сумма: {total:.2f}")
        self.refresh_items_panel()
        if hasattr(self.master, "refresh_orders"):
            self.master.refresh_orders()

    def print_order(self) -> None:
        if not self.order_id:
            messagebox.showwarning("Внимание", "Сначала сохраните заказ")
            return
        order = self.db.get_order_by_id(self.order_id)
        if not order:
            return
        services = self.db.get_order_services(self.order_id)
        rows_html = "".join(
            f"<tr><td>{s['service_name']}</td><td style='text-align:right'>{s['price']:.2f}</td><td style='text-align:center'>{s['quantity']}</td><td style='text-align:right'>{float(s['price']) * int(s['quantity']):.2f}</td></tr>"
            for s in services
        )
        html = f"""
<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>Заказ {order['order_number']}</title>
<style>body{{font-family:Arial,sans-serif;margin:20px}}table{{width:100%;border-collapse:collapse}}th,td{{padding:8px;border-bottom:1px solid #ddd}}th{{background:#222;color:#fff}}.total{{text-align:right;margin-top:16px;font-size:18px;font-weight:bold}}</style>
</head><body>
<h2>Заказ-наряд {order['order_number']}</h2>
<p>Клиент: {order['client_name']} | {order['phone']}</p>
<p>Дата: {order['created_date']}</p>
<table><thead><tr><th>Услуга</th><th style='text-align:right'>Цена</th><th style='text-align:center'>Кол-во</th><th style='text-align:right'>Сумма</th></tr></thead><tbody>{rows_html}</tbody></table>
<div class="total">ИТОГО: {order['total_sum']:.2f}</div>
</body></html>
"""
        with tempfile.NamedTemporaryFile(mode="w", suffix=".html", delete=False, encoding="utf-8") as f:
            f.write(html)
            path = f.name
        webbrowser.open(path)


class PeriodManager(ctk.CTkToplevel):
    def __init__(self, parent, db: Database, on_change):
        super().__init__(parent)
        self.db = db
        self.on_change = on_change
        self.title("Периоды цен")
        self.geometry("860x560")
        self.setup_ui()
        self.load_periods()

    def setup_ui(self) -> None:
        top = ctk.CTkFrame(self)
        top.pack(fill="x", padx=10, pady=10)
        ctk.CTkButton(top, text="Новый период", command=self.create_period, fg_color="green", width=140).pack(side="left", padx=5)
        ctk.CTkButton(top, text="Импорт Excel", command=self.import_from_excel, width=140).pack(side="left", padx=5)
        ctk.CTkButton(top, text="Экспорт активного", command=self.export_active_to_excel, width=160).pack(side="left", padx=5)
        ctk.CTkButton(top, text="Закрыть", command=self.destroy, fg_color="gray", width=120).pack(side="right", padx=5)
        self.list_frame = ctk.CTkScrollableFrame(self, height=440)
        self.list_frame.pack(fill="both", expand=True, padx=10, pady=5)

    def load_periods(self) -> None:
        for child in self.list_frame.winfo_children():
            child.destroy()
        for row in self.db.get_all_periods():
            frame = ctk.CTkFrame(self.list_frame, fg_color=("#1A3A1A", "#0A2A0A") if row["is_active"] else "transparent")
            frame.pack(fill="x", pady=1)
            ctk.CTkLabel(frame, text=str(row["id"]), width=60).pack(side="left", padx=5)
            ctk.CTkLabel(frame, text=row["name"], width=320, anchor="w").pack(side="left", padx=5)
            ctk.CTkLabel(frame, text=row["start_date"], width=170, anchor="w").pack(side="left", padx=5)
            ctk.CTkLabel(frame, text="Активен" if row["is_active"] else "Завершен", width=100).pack(side="left", padx=5)
            ctk.CTkButton(frame, text="Цены", width=80, command=lambda pid=row["id"]: self.show_prices(pid)).pack(side="left", padx=4)
            if not row["is_active"]:
                ctk.CTkButton(frame, text="Активировать", width=110, fg_color="blue", command=lambda pid=row["id"]: self.activate(pid)).pack(side="left", padx=4)

    def activate(self, period_id: int) -> None:
        self.db.activate_period(period_id)
        self.load_periods()
        self.on_change()

    def show_prices(self, period_id: int) -> None:
        win = ctk.CTkToplevel(self)
        win.title(f"Цены периода {period_id}")
        win.geometry("560x480")
        frame = ctk.CTkScrollableFrame(win, height=400)
        frame.pack(fill="both", expand=True, padx=10, pady=10)
        for idx, (name, price) in enumerate(self.db.get_period_prices(period_id), 1):
            row = ctk.CTkFrame(frame, fg_color="transparent")
            row.pack(fill="x", pady=1)
            ctk.CTkLabel(row, text=str(idx), width=45).pack(side="left", padx=4)
            ctk.CTkLabel(row, text=name, width=360, anchor="w").pack(side="left", padx=4)
            ctk.CTkLabel(row, text=f"{price:.2f}", width=100, anchor="e").pack(side="left", padx=4)

    def _collect_current_prices(self) -> list[tuple[str, float]]:
        return [(row["name"], float(row["price"])) for row in self.db.get_active_services()]

    def create_period(self) -> None:
        dialog = ctk.CTkInputDialog(text="Название нового периода", title="Новый период")
        name = dialog.get_input()
        if not name:
            return
        prices = self._collect_current_prices()
        if not prices:
            messagebox.showwarning("Ошибка", "Нет услуг для создания периода")
            return
        self.db.create_period_from_prices(name.strip(), prices)
        self.load_periods()
        self.on_change()

    def import_from_excel(self) -> None:
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not path:
            return
        wb = load_workbook(path)
        ws = wb.active
        prices: list[tuple[str, float]] = []
        for row in ws.iter_rows(min_row=2):
            if len(row) < 3:
                continue
            name = row[1].value
            value = row[2].value
            if not name or value is None:
                continue
            try:
                prices.append((str(name).strip(), float(str(value).replace(",", "."))))
            except ValueError:
                continue
        if not prices:
            messagebox.showwarning("Ошибка", "Не удалось прочитать цены из файла")
            return
        period_name = f"Период с {datetime.datetime.now().strftime('%d.%m.%Y')}"
        self.db.create_period_from_prices(period_name, prices)
        self.load_periods()
        self.on_change()

    def export_active_to_excel(self) -> None:
        prices = self.db.get_period_prices()
        if not prices:
            messagebox.showwarning("Ошибка", "Нет активного периода")
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Цены"
        ws.merge_cells("A1:C1")
        ws["A1"] = "Прайс-лист"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A2"] = f"Дата: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}"
        for col, text in enumerate(["№", "Услуга", "Цена"], 1):
            cell = ws.cell(row=4, column=col, value=text)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for idx, (name, price) in enumerate(prices, 1):
            ws.cell(row=4 + idx, column=1, value=idx)
            ws.cell(row=4 + idx, column=2, value=name)
            ws.cell(row=4 + idx, column=3, value=price)
        ws.column_dimensions[get_column_letter(1)].width = 8
        ws.column_dimensions[get_column_letter(2)].width = 52
        ws.column_dimensions[get_column_letter(3)].width = 18
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"prices_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        )
        if path:
            wb.save(path)
            messagebox.showinfo("Успех", "Файл сохранен")


class StatisticsWindow(ctk.CTkToplevel):
    def __init__(self, parent, db: Database):
        super().__init__(parent)
        self.db = db
        self.period = "week"
        self.title("Статистика")
        self.geometry("1020x680")
        self.setup_ui()
        self.load_statistics("week")

    def setup_ui(self) -> None:
        top = ctk.CTkFrame(self)
        top.pack(fill="x", padx=10, pady=10)
        for text, value in [("Неделя", "week"), ("Месяц", "month"), ("Год", "year")]:
            ctk.CTkButton(top, text=text, width=120, command=lambda v=value: self.load_statistics(v)).pack(side="left", padx=5)
        self.total_label = ctk.CTkLabel(top, text="Выручка: 0.00")
        self.total_label.pack(side="right", padx=8)
        self.orders_label = ctk.CTkLabel(top, text="Заказов: 0")
        self.orders_label.pack(side="right", padx=8)
        self.avg_label = ctk.CTkLabel(top, text="Средний чек: 0.00")
        self.avg_label.pack(side="right", padx=8)
        self.table = ctk.CTkScrollableFrame(self, height=560)
        self.table.pack(fill="both", expand=True, padx=10, pady=5)

    def load_statistics(self, period: str) -> None:
        self.period = period
        stats = self.db.get_statistics(period)
        self.total_label.configure(text=f"Выручка: {stats['total_sum']:.2f}")
        self.orders_label.configure(text=f"Заказов: {stats['total_orders']}")
        self.avg_label.configure(text=f"Средний чек: {stats['avg_check']:.2f}")
        for child in self.table.winfo_children():
            child.destroy()
        for row in stats["orders"]:
            frame = ctk.CTkFrame(self.table, fg_color="transparent")
            frame.pack(fill="x", pady=1)
            ctk.CTkLabel(frame, text=row["order_number"], width=130, anchor="w").pack(side="left", padx=5)
            ctk.CTkLabel(frame, text=row["client_name"], width=180, anchor="w").pack(side="left", padx=5)
            ctk.CTkLabel(frame, text=row["phone"], width=150, anchor="w").pack(side="left", padx=5)
            ctk.CTkLabel(frame, text=row["created_date"], width=150, anchor="w").pack(side="left", padx=5)
            ctk.CTkLabel(frame, text=f"{row['total_sum']:.2f}", width=110, anchor="e").pack(side="left", padx=5)


class MainApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.db = Database()
        self.orders: list[sqlite3.Row] = []
        self.selected_order_id: int | None = None
        self.title("ИТ-мастерская — Управление заказ-нарядами")
        self.geometry("1240x760")
        self.minsize(1040, 620)
        self.setup_ui()
        self.refresh_orders()

    def setup_ui(self) -> None:
        header = ctk.CTkFrame(self)
        header.pack(fill="x", padx=10, pady=8)
        ctk.CTkLabel(header, text="Управление заказ-нарядами", font=("Arial", 22, "bold")).pack(side="left", padx=8)
        tools = ctk.CTkFrame(header, fg_color="transparent")
        tools.pack(side="right", padx=8)
        ctk.CTkButton(tools, text="Периоды цен", width=130, fg_color="orange", command=self.open_period_manager).pack(side="left", padx=4)
        ctk.CTkButton(tools, text="Статистика", width=130, fg_color="purple", command=self.open_statistics).pack(side="left", padx=4)
        ctk.CTkButton(tools, text="Услуги", width=120, fg_color="blue", command=self.open_service_manager).pack(side="left", padx=4)
        ctk.CTkButton(tools, text="Новый заказ", width=130, fg_color="green", command=self.create_order).pack(side="left", padx=4)
        ctk.CTkButton(tools, text="Открыть", width=100, command=self.open_selected_order).pack(side="left", padx=4)
        ctk.CTkButton(tools, text="Удалить", width=100, fg_color="red", command=self.delete_selected_order).pack(side="left", padx=4)

        body = ctk.CTkFrame(self)
        body.pack(fill="both", expand=True, padx=10, pady=5)
        left = ctk.CTkFrame(body)
        left.pack(side="left", fill="both", expand=True, padx=(0, 5))
        right = ctk.CTkFrame(body, width=360)
        right.pack(side="left", fill="y", padx=(5, 0))

        search_frame = ctk.CTkFrame(left)
        search_frame.pack(fill="x", padx=5, pady=5)
        self.search_orders = ctk.CTkEntry(search_frame, placeholder_text="Поиск по номеру, клиенту, телефону")
        self.search_orders.pack(side="left", fill="x", expand=True, padx=(0, 5))
        self.search_orders.bind("<KeyRelease>", lambda _: self.display_orders())
        ctk.CTkButton(search_frame, text="Обновить", width=120, command=self.refresh_orders).pack(side="left")

        self.orders_panel = ctk.CTkScrollableFrame(left, height=610)
        self.orders_panel.pack(fill="both", expand=True, padx=5, pady=5)

        ctk.CTkLabel(right, text="Клиенты", font=("Arial", 16, "bold")).pack(pady=(10, 5))
        self.clients_panel = ctk.CTkScrollableFrame(right, height=500)
        self.clients_panel.pack(fill="both", expand=True, padx=8, pady=5)
        self.stats_label = ctk.CTkLabel(right, text="Заказов: 0\nАктивных: 0\nКлиентов: 0", justify="left")
        self.stats_label.pack(pady=10)

    def refresh_orders(self) -> None:
        self.orders = self.db.get_all_orders()
        self.display_orders()
        self.display_clients()
        self.update_stats()

    def display_orders(self) -> None:
        for child in self.orders_panel.winfo_children():
            child.destroy()
        query = self.search_orders.get().strip().lower()
        rows = self.orders
        if query:
            rows = [
                row
                for row in rows
                if query in row["order_number"].lower() or query in row["client_name"].lower() or query in row["phone"].lower()
            ]
        for row in rows:
            frame = ctk.CTkFrame(
                self.orders_panel,
                fg_color=("#3B8ED0", "#1F6AA5") if self.selected_order_id == row["id"] else "transparent",
            )
            frame.pack(fill="x", pady=1)
            ctk.CTkLabel(frame, text=row["order_number"], width=120, anchor="w").pack(side="left", padx=4)
            ctk.CTkLabel(frame, text=row["client_name"], width=180, anchor="w").pack(side="left", padx=4)
            ctk.CTkLabel(frame, text=row["phone"], width=150, anchor="w").pack(side="left", padx=4)
            ctk.CTkLabel(frame, text=row["created_date"], width=150, anchor="w").pack(side="left", padx=4)
            ctk.CTkLabel(frame, text=f"{row['total_sum']:.2f}", width=100, anchor="e").pack(side="left", padx=4)
            ctk.CTkLabel(frame, text=row["status"], width=90, anchor="center").pack(side="left", padx=4)
            ctk.CTkButton(frame, text="Открыть", width=90, command=lambda oid=row["id"]: self.open_order(oid)).pack(side="right", padx=4)
            frame.bind("<Button-1>", lambda _, oid=row["id"]: self.select_order(oid))

    def select_order(self, order_id: int) -> None:
        self.selected_order_id = order_id
        self.display_orders()

    def display_clients(self) -> None:
        for child in self.clients_panel.winfo_children():
            child.destroy()
        for row in self.db.get_all_clients():
            frame = ctk.CTkFrame(self.clients_panel, fg_color="transparent")
            frame.pack(fill="x", pady=1)
            ctk.CTkLabel(frame, text=row["name"], width=140, anchor="w").pack(side="left", padx=4)
            ctk.CTkLabel(frame, text=row["phone"], width=130, anchor="w").pack(side="left", padx=4)
            ctk.CTkLabel(frame, text=str(row["total_orders"]), width=50, anchor="center").pack(side="left", padx=4)
            ctk.CTkLabel(frame, text=f"{row['total_spent']:.2f}", width=90, anchor="e").pack(side="left", padx=4)

    def update_stats(self) -> None:
        total = len(self.orders)
        active = len([o for o in self.orders if o["status"] == "active"])
        clients = len(self.db.get_all_clients())
        self.stats_label.configure(text=f"Заказов: {total}\nАктивных: {active}\nКлиентов: {clients}")

    def create_order(self) -> None:
        selector = ClientSelector(self, self.db)
        selector.grab_set()
        self.wait_window(selector)
        if selector.selected_client_id:
            editor = OrderEditor(self, self.db, client_id=selector.selected_client_id)
            editor.grab_set()

    def open_order(self, order_id: int) -> None:
        editor = OrderEditor(self, self.db, order_id=order_id)
        editor.grab_set()

    def open_selected_order(self) -> None:
        if not self.selected_order_id:
            messagebox.showwarning("Внимание", "Заказ не выбран")
            return
        self.open_order(self.selected_order_id)

    def delete_selected_order(self) -> None:
        if not self.selected_order_id:
            messagebox.showwarning("Внимание", "Заказ не выбран")
            return
        if not messagebox.askyesno("Подтверждение", "Удалить выбранный заказ?"):
            return
        self.db.delete_order(self.selected_order_id)
        self.selected_order_id = None
        self.refresh_orders()

    def open_service_manager(self) -> None:
        manager = ServiceManager(self, self.db, self.refresh_orders)
        manager.grab_set()

    def open_statistics(self) -> None:
        win = StatisticsWindow(self, self.db)
        win.grab_set()

    def open_period_manager(self) -> None:
        win = PeriodManager(self, self.db, self.refresh_orders)
        win.grab_set()

    def on_close(self) -> None:
        self.db.close()
        self.destroy()


if __name__ == "__main__":
    app = MainApp()
    app.protocol("WM_DELETE_WINDOW", app.on_close)
    app.mainloop()

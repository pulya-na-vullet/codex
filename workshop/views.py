from __future__ import annotations

import json
import time
from datetime import timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.conf import settings
from django.contrib import messages
from django.db.models import Q
from django.http import HttpRequest, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_http_methods, require_POST

from workshop.audit import log_action
from workshop.models import (
    AcceptanceAct,
    AcceptanceActStatus,
    AuditLog,
    Client,
    DeviceType,
    Order,
    OrderLine,
    OrderStatus,
    PaymentMethod,
    Service,
    next_numbered,
)
from workshop.pdf import build_acceptance_act_pdf, build_order_pdf
from workshop.printing import PRINT_COPIES, enqueue_pdf_print
from workshop.services import build_service_catalog_tree, category_choices, ensure_category_path
from workshop.utils import normalize_rf_phone


def login_view(request: HttpRequest):
    next_url = request.GET.get("next") or request.POST.get("next") or "/"
    if request.session.get("workshop_authenticated"):
        return redirect(next_url)
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")
        if username == settings.WORKSHOP_USERNAME and password == settings.WORKSHOP_PASSWORD:
            request.session.flush()
            request.session["workshop_authenticated"] = True
            request.session["workshop_username"] = username
            request.session["workshop_last_active"] = time.time()
            log_action(request, "login", details=f"user={username}")
            messages.success(request, "Вход выполнен")
            return redirect(next_url)
        messages.error(request, "Неверный логин или пароль")
    return render(request, "workshop/login.html", {"next_url": next_url, "title": "Вход"})


@require_http_methods(["GET", "POST"])
def logout_view(request: HttpRequest):
    log_action(request, "logout")
    request.session.flush()
    messages.success(request, "Вы вышли из системы")
    return redirect("login")


def dashboard(request: HttpRequest):
    from workshop.models import AcceptanceActStatus

    orders_in_work = Order.objects.filter(status=OrderStatus.ACTIVE).count()
    diagnostics_in_work = AcceptanceAct.objects.filter(status=AcceptanceActStatus.DIAGNOSTICS).count()
    calls_needed = (
        Order.objects.filter(status=OrderStatus.READY_CALL).count()
        + AcceptanceAct.objects.filter(status=AcceptanceActStatus.DIAGNOSTICS_DONE).count()
    )
    return render(
        request,
        "workshop/dashboard.html",
        {
            "orders_in_work": orders_in_work,
            "diagnostics_in_work": diagnostics_in_work,
            "calls_needed": calls_needed,
        },
    )

def statistics(request: HttpRequest):
    from calendar import Calendar
    from collections import defaultdict

    period = request.GET.get("period", "week")
    if period not in {"week", "month", "year"}:
        period = "week"

    now = timezone.localtime()
    if period == "week":
        start = now - timedelta(days=7)
        end = now
        period_label = "за 7 дней"
    elif period == "year":
        start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        end = now
        period_label = f"за {now.year} год"
    else:
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end = now
        period_label = now.strftime("%m.%Y")

    orders = list(
        Order.objects.select_related("client").filter(created_at__gte=start, created_at__lte=end).order_by("-id")
    )
    acts = list(
        AcceptanceAct.objects.select_related("client").filter(created_at__gte=start, created_at__lte=end).order_by("-id")
    )
    total_sum = sum((o.total_sum for o in orders), Decimal("0"))
    total_orders = len(orders)
    paid_sum = sum((o.total_sum for o in orders if o.is_paid), Decimal("0"))
    paid_count = sum(1 for o in orders if o.is_paid)
    debt_sum = sum((o.total_sum for o in orders if o.is_debtor), Decimal("0"))
    debt_count = sum(1 for o in orders if o.is_debtor)
    in_progress_sum = sum((o.total_sum for o in orders if o.is_in_progress), Decimal("0"))
    in_progress_count = sum(1 for o in orders if o.is_in_progress)
    # Справочная сумма трёх групп (пересечения возможны: «в работе» может быть и долгом).
    breakdown_total = paid_sum + debt_sum + in_progress_sum
    avg_check = (paid_sum / paid_count) if paid_count else Decimal("0")

    # Unique contacted clients (orders + acceptance acts)
    visitors_map: dict[int, dict] = {}
    for o in orders:
        if not o.client_id:
            continue
        row = visitors_map.setdefault(
            o.client_id,
            {
                "id": o.client_id,
                "name": o.client.name,
                "phone": o.client.phone,
                "orders": 0,
                "acts": 0,
                "spent": Decimal("0"),
            },
        )
        row["orders"] += 1
        row["spent"] += o.total_sum
    for a in acts:
        row = visitors_map.setdefault(
            a.client_id,
            {
                "id": a.client_id,
                "name": a.client.name,
                "phone": a.client.phone,
                "orders": 0,
                "acts": 0,
                "spent": Decimal("0"),
            },
        )
        row["acts"] += 1
    visitors = sorted(visitors_map.values(), key=lambda r: (r["orders"] + r["acts"], r["spent"]), reverse=True)

    # Top clients by spend in period
    top_clients = sorted(visitors_map.values(), key=lambda r: r["spent"], reverse=True)[:10]

    # Monthly breakdown (year view / shared)
    monthly_map: dict[str, dict] = {}
    year_orders_qs = Order.objects.filter(created_at__year=now.year).only("created_at", "total_sum")
    if period == "year":
        source_orders = year_orders_qs
    else:
        source_orders = Order.objects.all().only("created_at", "total_sum")
    for o in source_orders:
        local = timezone.localtime(o.created_at)
        key = local.strftime("%Y-%m")
        label = local.strftime("%m.%Y")
        bucket = monthly_map.setdefault(key, {"key": key, "label": label, "orders": 0, "revenue": Decimal("0")})
        bucket["orders"] += 1
        bucket["revenue"] += o.total_sum
    monthly = [monthly_map[k] for k in sorted(monthly_map.keys(), reverse=True)[:12]]

    # Calendar heatmap for current month
    calendar_weeks = []
    max_day_visits = 0
    if period == "month":
        day_counts: dict = defaultdict(int)
        month_start = start
        if now.month == 12:
            month_end = now.replace(year=now.year + 1, month=1, day=1)
        else:
            month_end = now.replace(month=now.month + 1, day=1)
        for o in Order.objects.filter(created_at__gte=month_start, created_at__lt=month_end).only("created_at"):
            day_counts[timezone.localtime(o.created_at).date()] += 1
        for a in AcceptanceAct.objects.filter(created_at__gte=month_start, created_at__lt=month_end).only("created_at"):
            day_counts[timezone.localtime(a.created_at).date()] += 1
        max_day_visits = max(day_counts.values()) if day_counts else 0
        cal = Calendar(firstweekday=0)  # Monday
        for week in cal.monthdatescalendar(now.year, now.month):
            cells = []
            for d in week:
                count = day_counts.get(d, 0)
                intensity = 0
                if max_day_visits and count:
                    intensity = max(1, round(count / max_day_visits * 4))
                cells.append(
                    {
                        "day": d.day,
                        "date": d,
                        "count": count,
                        "in_month": d.month == now.month,
                        "intensity": intensity,
                        "is_today": d == now.date(),
                    }
                )
            calendar_weeks.append(cells)

    # Year-over-year comparison (for year period; scaffold for the future)
    yoy = None
    if period == "year":
        prev_start = start.replace(year=start.year - 1)
        prev_end = start
        prev_orders = list(Order.objects.filter(created_at__gte=prev_start, created_at__lt=prev_end))
        prev_sum = sum((o.total_sum for o in prev_orders), Decimal("0"))
        prev_count = len(prev_orders)
        prev_avg = (prev_sum / prev_count) if prev_count else Decimal("0")

        def _delta(curr, prev):
            if prev == 0:
                return None if curr == 0 else Decimal("100")
            return ((curr - prev) / prev * Decimal("100")).quantize(Decimal("0.1"))

        yoy = {
            "prev_year": start.year - 1,
            "curr_year": start.year,
            "prev_orders": prev_count,
            "prev_sum": prev_sum,
            "prev_avg": prev_avg,
            "orders_delta_pct": _delta(Decimal(total_orders), Decimal(prev_count)),
            "sum_delta_pct": _delta(total_sum, prev_sum),
            "avg_delta_pct": _delta(avg_check, prev_avg),
        }

    return render(
        request,
        "workshop/statistics.html",
        {
            "period": period,
            "period_label": period_label,
            "stats": {
                "total_sum": total_sum,
                "total_orders": total_orders,
                "avg_check": avg_check,
                "visitors_count": len(visitors),
                "acts_count": len(acts),
                "paid_sum": paid_sum,
                "paid_count": paid_count,
                "debt_sum": debt_sum,
                "debt_count": debt_count,
                "in_progress_sum": in_progress_sum,
                "in_progress_count": in_progress_count,
                "breakdown_total": breakdown_total,
            },
            "orders": orders,
            "visitors": visitors,
            "monthly": monthly,
            "top_clients": top_clients,
            "calendar_weeks": calendar_weeks,
            "max_day_visits": max_day_visits,
            "yoy": yoy,
            "weekday_labels": ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"],
            "order_statuses": OrderStatus.choices,
        },
    )


def work_queue(request: HttpRequest):
    orders = list(
        Order.objects.select_related("client")
        .filter(status=OrderStatus.ACTIVE)
        .order_by("created_at", "id")
    )
    acts = list(
        AcceptanceAct.objects.select_related("client", "order")
        .filter(status=AcceptanceActStatus.DIAGNOSTICS)
        .order_by("created_at", "id")
    )
    call_orders = list(
        Order.objects.select_related("client")
        .filter(status=OrderStatus.READY_CALL)
        .order_by("closed_at", "id")
    )
    call_acts = list(
        AcceptanceAct.objects.select_related("client", "order")
        .filter(status=AcceptanceActStatus.DIAGNOSTICS_DONE)
        .order_by("finished_at", "id")
    )
    return render(
        request,
        "workshop/work_queue.html",
        {
            "orders": orders,
            "acts": acts,
            "call_orders": call_orders,
            "call_acts": call_acts,
            "order_statuses": OrderStatus.choices,
            "act_statuses": AcceptanceActStatus.choices,
            "count": len(orders) + len(acts),
            "call_count": len(call_orders) + len(call_acts),
        },
    )


@require_POST
def order_set_status(request: HttpRequest, order_id: int):
    from workshop.messaging import maybe_notify_order_done

    order = get_object_or_404(Order.objects.select_related("client"), pk=order_id)
    status = request.POST.get("status", "").strip()
    if status not in dict(OrderStatus.choices):
        messages.warning(request, "Некорректный статус")
        return redirect(request.POST.get("next") or "work_queue")
    old = order.status
    username = str(request.session.get("workshop_username") or "")
    order.apply_status(status)
    maybe_notify_order_done(order, old_status=old, username=username)
    log_action(
        request,
        "order_set_status",
        entity_type="order",
        entity_id=order.id,
        details=f"{order.order_number}: {old} → {order.status}",
    )
    messages.success(request, f"Статус заказа {order.order_number}: {order.get_status_display()}")
    next_url = request.POST.get("next") or ""
    if next_url.startswith("/"):
        return redirect(next_url)
    if order.status in {OrderStatus.ACTIVE, OrderStatus.READY_CALL}:
        return redirect("work_queue")
    return redirect("order_detail", order_id=order.id)


@require_POST
def order_mark_called(request: HttpRequest, order_id: int):
    from workshop.messaging import maybe_notify_order_done

    order = get_object_or_404(Order.objects.select_related("client"), pk=order_id)
    old = order.status
    username = str(request.session.get("workshop_username") or "")
    order.mark_client_called()
    maybe_notify_order_done(order, old_status=old, username=username)
    log_action(
        request,
        "order_client_called",
        entity_type="order",
        entity_id=order.id,
        details=order.order_number,
    )
    messages.success(request, f"Звонок по заказу {order.order_number} отмечен — статус «Выполнена»")
    next_url = request.POST.get("next") or ""
    if next_url.startswith("/"):
        return redirect(next_url)
    return redirect("work_queue")


@require_POST
def acceptance_set_status(request: HttpRequest, act_id: int):
    from workshop.messaging import maybe_notify_diagnostics_done

    act = get_object_or_404(AcceptanceAct.objects.select_related("client"), pk=act_id)
    status = request.POST.get("status", "").strip()
    if status not in dict(AcceptanceActStatus.choices):
        messages.warning(request, "Некорректный статус акта")
        return redirect(request.POST.get("next") or "work_queue")
    old = act.status
    username = str(request.session.get("workshop_username") or "")
    act.apply_status(status)
    maybe_notify_diagnostics_done(act, old_status=old, username=username)
    log_action(
        request,
        "acceptance_set_status",
        entity_type="acceptance",
        entity_id=act.id,
        details=f"{act.act_number}: {old} → {act.status}",
    )
    messages.success(request, f"Статус акта {act.act_number}: {act.get_status_display()}")
    next_url = request.POST.get("next") or ""
    if next_url.startswith("/"):
        return redirect(next_url)
    return redirect("work_queue")


@require_POST
def acceptance_mark_called(request: HttpRequest, act_id: int):
    act = get_object_or_404(AcceptanceAct, pk=act_id)
    act.mark_client_called()
    log_action(
        request,
        "acceptance_client_called",
        entity_type="acceptance",
        entity_id=act.id,
        details=act.act_number,
    )
    messages.success(request, f"Звонок по акту {act.act_number} отмечен — статус «Выполнена»")
    next_url = request.POST.get("next") or ""
    if next_url.startswith("/"):
        return redirect(next_url)
    return redirect("work_queue")


@require_http_methods(["GET", "POST"])
def clients(request: HttpRequest):
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        phone_raw = request.POST.get("phone", "").strip()
        comment = request.POST.get("comment", "").strip()
        normalized = normalize_rf_phone(phone_raw)
        if not name:
            messages.warning(request, "Введите имя клиента")
        elif not normalized:
            messages.warning(request, "Введите корректный номер РФ")
        elif Client.objects.filter(phone=normalized).exists():
            messages.warning(request, "Клиент с таким телефоном уже существует")
        else:
            Client.objects.create(name=name, phone=normalized, comment=comment)
            log_action(request, "client_create", entity_type="client", details=f"{name} {normalized}")
            messages.success(request, "Клиент успешно добавлен")
        return redirect("clients")

    query = request.GET.get("q", "").strip()
    qs = Client.objects.all()
    if query:
        qs = qs.filter(Q(name__icontains=query) | Q(phone__icontains=query) | Q(comment__icontains=query))
    rows = list(qs)
    enriched = []
    for c in rows:
        enriched.append(
            {
                "id": c.id,
                "name": c.name,
                "phone": c.phone,
                "comment": c.comment,
                "total_orders": c.total_orders,
                "total_spent": c.total_spent,
                "is_regular": c.is_regular,
                "discount_percent": c.discount_percent,
            }
        )
    return render(request, "workshop/clients.html", {"clients": enriched, "query": query})


@require_http_methods(["GET", "POST"])
def client_detail(request: HttpRequest, client_id: int):
    client = get_object_or_404(Client, pk=client_id)
    if request.method == "POST":
        comment = request.POST.get("comment", "").strip()
        max_user_id = request.POST.get("max_user_id", "").strip()
        allow_marketing = request.POST.get("allow_marketing_sms") == "1"
        discount_raw = request.POST.get("discount_percent", "0").strip().replace(",", ".")
        try:
            discount_value = Decimal(discount_raw or "0")
        except (InvalidOperation, ValueError):
            messages.warning(request, "Некорректная скидка")
            return redirect("client_detail", client_id=client.id)
        client.comment = comment
        client.max_user_id = max_user_id
        client.allow_marketing_sms = allow_marketing
        client.set_discount_percent(discount_value, manual=True, save=False)
        client.save(update_fields=["comment", "max_user_id", "allow_marketing_sms", "discount_percent", "discount_manual"])
        log_action(
            request,
            "client_update",
            entity_type="client",
            entity_id=client.id,
            details=(
                f"{client.name}: max={max_user_id or '-'} marketing={allow_marketing} "
                f"discount={client.discount_percent}%"
            ),
        )
        messages.success(request, "Данные клиента сохранены")
        return redirect("client_detail", client_id=client.id)

    client.apply_auto_regular_discount(save=True)
    orders = client.orders.all()
    acts = client.acceptance_acts.all()[:20]
    ctx_client = {
        "id": client.id,
        "name": client.name,
        "phone": client.phone,
        "comment": client.comment,
        "max_user_id": client.max_user_id,
        "allow_marketing_sms": client.allow_marketing_sms,
        "total_orders": client.total_orders,
        "total_spent": client.total_spent,
        "is_regular": client.is_regular,
        "discount_percent": client.discount_percent,
        "discount_manual": client.discount_manual,
    }
    return render(
        request,
        "workshop/client_detail.html",
        {"client": ctx_client, "orders": orders, "acts": acts},
    )


@require_POST
def client_delete(request: HttpRequest, client_id: int):
    client = get_object_or_404(Client, pk=client_id)
    if client.orders.exists() or client.acceptance_acts.exists():
        messages.warning(
            request,
            "Нельзя удалить клиента: есть связанные заказ-наряды или акты. Сначала удалите их.",
        )
        return redirect("client_detail", client_id=client_id)
    details = f"{client.name} {client.phone}"
    client.delete()
    log_action(request, "client_delete", entity_type="client", entity_id=client_id, details=details)
    messages.success(request, "Клиент удалён")
    return redirect("clients")


@require_GET
def export_clients_excel(request: HttpRequest):
    try:
        from openpyxl import Workbook
    except ImportError:
        messages.warning(request, "Установите openpyxl: pip install openpyxl")
        return redirect("clients")
    wb = Workbook()
    ws = wb.active
    ws.title = "Клиенты"
    ws.append(["номер телефона", "имя", "комментарий"])
    for c in Client.objects.order_by("id"):
        ws.append([c.phone, c.name, c.comment])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="clients.xlsx"'
    return response


@require_POST
def import_clients_excel(request: HttpRequest):
    try:
        from openpyxl import load_workbook
    except ImportError:
        messages.warning(request, "Установите openpyxl: pip install openpyxl")
        return redirect("clients")
    upload = request.FILES.get("excel_file")
    if not upload:
        messages.warning(request, "Выберите Excel-файл")
        return redirect("clients")
    wb = load_workbook(upload, read_only=True, data_only=True)
    ws = wb.active
    imported = skipped_existing = skipped_invalid = 0
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row:
            continue
        values = list(row) + [None, None, None]
        phone = "" if values[0] is None else str(values[0]).strip()
        name = "" if values[1] is None else str(values[1]).strip()
        comment = "" if values[2] is None else str(values[2]).strip()
        if idx == 1 and ("телефон" in phone.lower() or phone.lower() in {"phone", "номер телефона"}):
            continue
        if not phone and not name:
            continue
        if not phone or not name:
            skipped_invalid += 1
            continue
        normalized = normalize_rf_phone(phone)
        if not normalized:
            skipped_invalid += 1
            continue
        if Client.objects.filter(phone=normalized).exists():
            skipped_existing += 1
            continue
        Client.objects.create(name=name, phone=normalized, comment=comment)
        imported += 1
    messages.success(
        request,
        f"Импорт завершён: добавлено {imported}, пропущено существующих {skipped_existing}, некорректных {skipped_invalid}",
    )
    return redirect("clients")


@require_http_methods(["GET", "POST"])
def services(request: HttpRequest):
    status_filter = request.GET.get("status", "active")
    if status_filter not in {"all", "active", "inactive"}:
        status_filter = "active"

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        price_raw = request.POST.get("price", "").strip().replace(",", ".")
        category = request.POST.get("category", "").strip() or "Основные"
        if not name:
            messages.warning(request, "Введите название услуги")
        else:
            try:
                price = Decimal(price_raw)
            except (InvalidOperation, ValueError):
                messages.warning(request, "Введите корректную цену")
                return redirect(f"/services?status={status_filter}")
            if price < 0:
                messages.warning(request, "Цена не может быть отрицательной")
            elif Service.objects.filter(name=name).exists():
                messages.warning(request, "Услуга с таким названием уже существует")
            else:
                cat = ensure_category_path(category)
                Service.objects.create(name=name, price=price, category=cat, is_active=True)
                log_action(request, "service_create", entity_type="service", details=f"{name}={price}")
                messages.success(request, "Услуга добавлена")
        return redirect(f"/services?status={status_filter}")

    qs = Service.objects.select_related("category")
    if status_filter == "active":
        qs = qs.filter(is_active=True)
    elif status_filter == "inactive":
        qs = qs.filter(is_active=False)

    return render(
        request,
        "workshop/services.html",
        {
            "services": qs,
            "categories": category_choices(),
            "status_filter": status_filter,
        },
    )


def _services_status_redirect(request: HttpRequest) -> str:
    status = request.POST.get("status") or request.GET.get("status") or "active"
    if status not in {"all", "active", "inactive"}:
        status = "active"
    return f"/services?status={status}"


def _price_list_groups(services) -> list[dict]:
    groups: list[dict] = []
    by_label: dict[str, list] = {}
    for service in services:
        label = service.category.path_label if service.category_id else "Основные"
        if label not in by_label:
            by_label[label] = []
            groups.append({"category": label, "services": by_label[label]})
        by_label[label].append(service)
    return groups


@require_GET
def services_print(request: HttpRequest):
    """Печатная форма прайс-листа активных услуг (две колонки)."""
    published_at = timezone.localdate()
    services = list(
        Service.objects.filter(is_active=True)
        .select_related("category")
        .order_by("category__name", "name")
    )
    mid = (len(services) + 1) // 2
    columns = [_price_list_groups(services[:mid]), _price_list_groups(services[mid:])]

    log_action(
        request,
        "services_price_list_print",
        entity_type="service",
        details=f"items={len(services)} date={published_at.isoformat()}",
    )
    return render(
        request,
        "workshop/print_services_price.html",
        {
            "columns": columns,
            "services_count": len(services),
            "published_at": published_at,
            "company_name": settings.COMPANY_NAME,
            "company_phone": settings.COMPANY_PHONE,
            "quality_phone": settings.QUALITY_PHONE,
            "company_address": settings.COMPANY_ADDRESS,
        },
    )


@require_POST
def service_delete(request: HttpRequest, service_id: int):
    service = get_object_or_404(Service, pk=service_id)
    details = service.name
    service.delete()
    log_action(request, "service_delete", entity_type="service", entity_id=service_id, details=details)
    messages.success(request, "Услуга удалена")
    return redirect(_services_status_redirect(request))


@require_POST
def service_toggle_active(request: HttpRequest, service_id: int):
    service = get_object_or_404(Service, pk=service_id)
    service.is_active = not service.is_active
    service.save(update_fields=["is_active"])
    status_label = "активна" if service.is_active else "неактивна"
    log_action(
        request,
        "service_toggle_active",
        entity_type="service",
        entity_id=service.id,
        details=f"{service.name}: {status_label}",
    )
    messages.success(request, f"Услуга «{service.name}» теперь {status_label}")
    return redirect(_services_status_redirect(request))


def orders_list(request: HttpRequest):
    return render(
        request,
        "workshop/orders.html",
        {"orders": Order.objects.select_related("client")},
    )


@require_GET
def orders_export_excel(request: HttpRequest):
    try:
        from openpyxl import Workbook
    except ImportError:
        messages.warning(request, "Установите openpyxl: pip install openpyxl")
        return redirect("orders")

    wb = Workbook()
    ws = wb.active
    ws.title = "Заказы"
    ws.append(
        [
            "Номер",
            "Клиент",
            "Телефон",
            "Дата",
            "Статус",
            "Сумма",
            "Оплата",
            "Мой налог",
            "Закрыт",
            "Звонок клиенту",
        ]
    )
    for order in Order.objects.select_related("client").order_by("-id"):
        ws.append(
            [
                order.order_number,
                order.client.name if order.client else "",
                order.client.phone if order.client else "",
                timezone.localtime(order.created_at).strftime("%d.%m.%Y %H:%M") if order.created_at else "",
                order.get_status_display(),
                float(order.total_sum or 0),
                order.get_payment_method_display(),
                "да" if order.mytax_issued else "нет",
                timezone.localtime(order.closed_at).strftime("%d.%m.%Y %H:%M") if order.closed_at else "",
                timezone.localtime(order.client_called_at).strftime("%d.%m.%Y %H:%M") if order.client_called_at else "",
            ]
        )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = timezone.localdate().isoformat()
    filename = f"orders + {stamp}.xlsx"
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    log_action(request, "orders_export_excel", entity_type="order", details=filename)
    return response


@require_http_methods(["GET", "POST"])
def order_create(request: HttpRequest):
    if request.method == "POST":
        client_id_raw = request.POST.get("client_id", "").strip()
        client = None
        if client_id_raw:
            client = get_object_or_404(Client, pk=int(client_id_raw))
        discount = client.discount_percent if client else Decimal("0")
        order = Order.objects.create(
            order_number=next_numbered("ORD", Order, "order_number"),
            client=client,
            discount_percent=discount,
        )
        if client:
            client.apply_auto_regular_discount(save=True)
            client.refresh_from_db(fields=["discount_percent"])
            order.discount_percent = client.discount_percent
            order.save(update_fields=["discount_percent"])
        log_action(
            request,
            "order_create",
            entity_type="order",
            entity_id=order.id,
            details=order.order_number,
        )
        messages.success(request, "Заказ создан")
        return redirect("order_detail", order_id=order.id)

    selected = request.GET.get("client_id")
    clients_qs = Client.objects.all()
    clients_data = [
        {
            "id": c.id,
            "name": c.name,
            "phone": c.phone,
            "is_regular": c.is_regular,
            "discount_percent": c.discount_percent,
        }
        for c in clients_qs
    ]
    return render(
        request,
        "workshop/order_new.html",
        {
            "clients": clients_data,
            "selected_client_id": int(selected) if selected and selected.isdigit() else None,
        },
    )


def order_detail(request: HttpRequest, order_id: int):
    order = get_object_or_404(Order.objects.select_related("client"), pk=order_id)
    client_ctx = None
    if order.client:
        c = order.client
        client_ctx = {
            "id": c.id,
            "is_regular": c.is_regular,
            "discount_percent": c.discount_percent,
            "total_orders": c.total_orders,
        }
    return render(
        request,
        "workshop/order_detail.html",
        {
            "order": order,
            "client": client_ctx,
            "lines": order.lines.all(),
            "catalog_tree": build_service_catalog_tree(active_only=True),
            "device_types": [c[0] for c in DeviceType.choices],
            "payment_methods": PaymentMethod.choices,
            "order_statuses": OrderStatus.choices,
        },
    )


@require_POST
def order_update_meta(request: HttpRequest, order_id: int):
    order = get_object_or_404(Order, pk=order_id)
    device_type = request.POST.get("device_type", DeviceType.PC)
    if device_type not in dict(DeviceType.choices):
        device_type = DeviceType.PC
    order.device_type = device_type
    order.extra_periphery = request.POST.get("extra_periphery", "").strip()
    # Preserve real newlines from textarea (including Shift+Enter as \n)
    order.technical_notes = (request.POST.get("technical_notes", "") or "").replace("\r\n", "\n").replace("\r", "\n")
    order.save(update_fields=["device_type", "extra_periphery", "technical_notes"])
    log_action(request, "order_update_meta", entity_type="order", entity_id=order.id, details=order.order_number)
    messages.success(request, "Данные заказ-наряда сохранены")
    return redirect("order_detail", order_id=order.id)


@require_POST
def order_add_service(request: HttpRequest, order_id: int):
    order = get_object_or_404(Order, pk=order_id)
    service_name = request.POST.get("service_name", "").strip()
    try:
        qty = max(1, int(request.POST.get("quantity", "1")))
    except ValueError:
        qty = 1
    service = Service.objects.filter(name=service_name, is_active=True).first()
    if not service:
        messages.warning(request, "Услуга не найдена")
        return redirect("order_detail", order_id=order.id)
    OrderLine.objects.create(
        order=order,
        service=service,
        service_name=service.name,
        unit_price=service.price,
        quantity=qty,
    )
    order.recalculate_totals()
    log_action(
        request,
        "order_add_service",
        entity_type="order",
        entity_id=order.id,
        details=f"{order.order_number}: {service.name} x{qty}",
    )
    messages.success(request, "Услуга добавлена в заказ")
    return redirect("order_detail", order_id=order.id)


@require_POST
def order_line_delete(request: HttpRequest, order_id: int, line_id: int):
    order = get_object_or_404(Order, pk=order_id)
    OrderLine.objects.filter(pk=line_id, order=order).delete()
    order.recalculate_totals()
    log_action(request, "order_line_delete", entity_type="order", entity_id=order.id, details=f"line={line_id}")
    messages.success(request, "Строка удалена")
    return redirect("order_detail", order_id=order.id)


@require_POST
def order_delete(request: HttpRequest, order_id: int):
    order = get_object_or_404(Order, pk=order_id)
    details = order.order_number
    order.delete()
    log_action(request, "order_delete", entity_type="order", entity_id=order_id, details=details)
    messages.success(request, "Заказ удалён")
    return redirect("orders")


@require_POST
def order_set_payment(request: HttpRequest, order_id: int):
    order = get_object_or_404(Order, pk=order_id)
    method = request.POST.get("payment_method", PaymentMethod.UNPAID)
    if method not in dict(PaymentMethod.choices):
        messages.warning(request, "Некорректный способ оплаты")
        return redirect("order_detail", order_id=order.id)

    note = request.POST.get("payment_note", "").strip()
    receipt = request.FILES.get("payment_receipt")

    if method == PaymentMethod.TRANSFER and not receipt and not order.payment_receipt:
        messages.warning(request, "Для оплаты переводом приложите скриншот чека")
        return redirect("order_detail", order_id=order.id)

    order.payment_method = method
    order.payment_note = note
    if method == PaymentMethod.UNPAID:
        order.payment_at = None
        if "clear_receipt" in request.POST:
            order.payment_receipt = None
    else:
        order.payment_at = timezone.now()
        if receipt:
            order.payment_receipt = receipt
    order.save()
    log_action(
        request,
        "order_payment",
        entity_type="order",
        entity_id=order.id,
        details=f"{order.order_number} method={method}",
    )
    messages.success(request, "Статус оплаты сохранён")
    return redirect("order_detail", order_id=order.id)


@require_POST
def order_set_mytax(request: HttpRequest, order_id: int):
    order = get_object_or_404(Order, pk=order_id)
    issued = request.POST.get("mytax_issued") == "1"
    receipt = request.FILES.get("mytax_receipt")
    order.mytax_issued = issued
    if issued:
        order.mytax_at = timezone.now()
        if receipt:
            order.mytax_receipt = receipt
    else:
        order.mytax_at = None
        if "clear_mytax_receipt" in request.POST:
            order.mytax_receipt = None
    order.save()
    log_action(
        request,
        "order_mytax",
        entity_type="order",
        entity_id=order.id,
        details=f"{order.order_number} issued={issued}",
    )
    messages.success(request, "Статус чека «Мой налог» сохранён")
    return redirect("order_detail", order_id=order.id)


def debtors_list(request: HttpRequest):
    from workshop.models import SmsSettings, debt_tracking_start, debtor_orders_queryset

    orders = list(debtor_orders_queryset().select_related("client").order_by("-debt_closed_at", "-id"))
    total_debt = sum((o.total_sum for o in orders), Decimal("0"))
    return render(
        request,
        "workshop/debtors.html",
        {
            "orders": orders,
            "total_debt": total_debt,
            "count": len(orders),
            "debt_tracking_start": debt_tracking_start().date(),
            "msg_settings": SmsSettings.get_solo(),
        },
    )


@require_POST
def debtors_sms_all(request: HttpRequest):
    from workshop.models import debtor_orders_queryset
    from workshop.messaging import send_debt_message_for_order

    username = str(request.session.get("workshop_username") or "")
    orders = list(debtor_orders_queryset().select_related("client"))
    ok = fail = 0
    for order in orders:
        if not order.client_id:
            fail += 1
            continue
        result = send_debt_message_for_order(order, username=username)
        if result.success:
            ok += 1
        else:
            fail += 1
    log_action(request, "max_debt_bulk", entity_type="messaging", details=f"ok={ok} fail={fail}")
    if ok:
        messages.success(request, f"Max: отправлено {ok}, ошибок {fail}")
    else:
        messages.warning(request, f"Сообщения не отправлены (ошибок: {fail}). Проверьте админ-панель Max.")
    return redirect("debtors")


@require_POST
def debtors_sms_one(request: HttpRequest, order_id: int):
    from workshop.messaging import send_debt_message_for_order

    order = get_object_or_404(Order.objects.select_related("client"), pk=order_id)
    username = str(request.session.get("workshop_username") or "")
    result = send_debt_message_for_order(order, username=username)
    log_action(
        request,
        "max_debt_one",
        entity_type="order",
        entity_id=order.id,
        details=f"{order.order_number}: {result.response[:200]}",
    )
    if result.success:
        note = " (симуляция)" if result.simulated else ""
        messages.success(request, f"Сообщение о долге отправлено в Max{note}: {order.order_number}")
    else:
        messages.warning(request, f"Не удалось отправить в Max: {result.response}")
    return redirect("debtors")


@require_http_methods(["GET", "POST"])
def admin_panel(request: HttpRequest):
    from workshop.models import SmsLog, SmsProvider, SmsSettings, YandexAiSettings
    from workshop.messaging import start_max_long_poll_worker
    from workshop.yandex_ai import run_daily_ai_report, start_ai_report_scheduler

    cfg = SmsSettings.get_solo()
    ai_cfg = YandexAiSettings.get_solo()

    if request.method == "POST":
        section = request.POST.get("section", "max").strip()
        if section == "ai_report_now":
            result = run_daily_ai_report(force=True)
            log_action(
                request,
                "ai_report_manual",
                entity_type="ai",
                details=f"ok={result.get('ok')} source={result.get('source')} {result.get('detail')}",
            )
            if result.get("ok"):
                messages.success(request, f"AI-отчёт отправлен ({result.get('source')})")
            else:
                messages.warning(request, f"AI-отчёт не отправлен: {result.get('detail')}")
            return redirect("admin_panel")

        if section == "ai":
            ai_cfg.enabled = request.POST.get("ai_enabled") == "1"
            ai_cfg.api_key = request.POST.get("api_key", "").strip()
            ai_cfg.folder_id = request.POST.get("folder_id", "").strip()
            ai_cfg.model_name = request.POST.get("model_name", "").strip() or "yandexgpt-lite"
            ai_cfg.admin_phone = request.POST.get("admin_phone", "").strip()
            ai_cfg.admin_max_user_id = request.POST.get("admin_max_user_id", "").strip()
            try:
                hour = int(request.POST.get("report_hour_msk", "20") or 20)
            except ValueError:
                hour = 20
            ai_cfg.report_hour_msk = max(0, min(23, hour))
            ai_cfg.save()
            start_ai_report_scheduler()
            log_action(request, "ai_settings_update", entity_type="ai", details=f"enabled={ai_cfg.enabled}")
            messages.success(request, "Настройки Яндекс ИИ сохранены")
            return redirect("admin_panel")

        cfg.enabled = request.POST.get("enabled") == "1"
        cfg.marketing_enabled = request.POST.get("marketing_enabled") == "1"
        cfg.long_poll_enabled = request.POST.get("long_poll_enabled") == "1"
        provider = request.POST.get("provider", SmsProvider.LOG_ONLY)
        if provider in dict(SmsProvider.choices):
            cfg.provider = provider
        cfg.bot_token = request.POST.get("bot_token", "").strip()
        cfg.bot_username = request.POST.get("bot_username", "").strip().lstrip("@")
        cfg.bot_link = request.POST.get("bot_link", "").strip()
        if not cfg.bot_link and cfg.bot_username:
            cfg.bot_link = f"https://max.ru/{cfg.bot_username}"
        cfg.welcome_text = request.POST.get("welcome_text", "").strip() or cfg.welcome_text
        cfg.debt_template = request.POST.get("debt_template", "").strip() or cfg.debt_template
        cfg.order_done_template = (
            request.POST.get("order_done_template", "").strip() or cfg.order_done_template
        )
        cfg.diagnostics_done_template = (
            request.POST.get("diagnostics_done_template", "").strip() or cfg.diagnostics_done_template
        )
        cfg.marketing_default_text = (
            request.POST.get("marketing_default_text", "").strip() or cfg.marketing_default_text
        )
        cfg.save()
        start_max_long_poll_worker()
        log_action(request, "max_settings_update", entity_type="messaging", details=f"provider={cfg.provider}")
        messages.success(request, "Настройки Max сохранены")
        return redirect("admin_panel")

    return render(
        request,
        "workshop/admin_panel.html",
        {
            "cfg": cfg,
            "ai_cfg": ai_cfg,
            "providers": SmsProvider.choices,
            "recent_messages": SmsLog.objects.select_related("client", "order")[:50],
        },
    )


@require_http_methods(["GET", "POST"])
def marketing_sms(request: HttpRequest):
    from django.db.models import Case, Count, IntegerField, Prefetch, Value, When

    from workshop.models import MarketingBlast, SmsKind, SmsLog, SmsSettings, debtor_orders_queryset
    from workshop.messaging import send_marketing_message

    cfg = SmsSettings.get_solo()
    debtor_client_ids = set(
        debtor_orders_queryset().exclude(client_id=None).values_list("client_id", flat=True)
    )

    if request.method == "POST":
        text = request.POST.get("text", "").strip()
        selected = request.POST.getlist("client_ids")
        username = str(request.session.get("workshop_username") or "")
        if not text:
            messages.warning(request, "Введите текст сообщения")
            return redirect("marketing_sms")
        if not selected:
            messages.warning(request, "Выберите хотя бы одного клиента")
            return redirect("marketing_sms")
        blast = MarketingBlast.objects.create(template_text=text, username=username)
        ok = fail = 0
        for cid in selected:
            client = Client.objects.filter(pk=cid).first()
            if not client:
                fail += 1
                continue
            result = send_marketing_message(client, text, username=username, blast=blast)
            if result.success:
                ok += 1
            else:
                fail += 1
        blast_id = blast.id
        if not blast.logs.exists():
            blast.delete()
            blast_id = None
        log_action(
            request,
            "max_marketing_bulk",
            entity_type="messaging",
            entity_id=blast_id,
            details=f"blast={blast_id or '-'} ok={ok} fail={fail}",
        )
        if ok:
            messages.success(request, f"Маркетинг Max: отправлено {ok}, ошибок {fail}")
        else:
            messages.warning(request, f"Сообщения не отправлены (ошибок: {fail}). Нужен Max user_id у клиента.")
        return redirect("marketing_sms")

    query = request.GET.get("q", "").strip()
    sort = request.GET.get("sort", "max").strip() or "max"
    allowed_sorts = {
        "name": ("name", "id"),
        "-name": ("-name", "id"),
        "date": ("created_at", "id"),
        "-date": ("-created_at", "id"),
        "max": ("-has_max", "name", "id"),
        "-max": ("has_max", "name", "id"),
        "regular": ("-orders_count", "name", "id"),
        "-regular": ("orders_count", "name", "id"),
    }
    order_by = allowed_sorts.get(sort, allowed_sorts["max"])

    qs = (
        Client.objects.exclude(id__in=debtor_client_ids)
        .filter(allow_marketing_sms=True)
        .annotate(
            orders_count=Count("orders"),
            has_max=Case(
                When(~Q(max_user_id=""), then=Value(1)),
                default=Value(0),
                output_field=IntegerField(),
            ),
        )
    )
    if query:
        qs = qs.filter(Q(name__icontains=query) | Q(phone__icontains=query) | Q(comment__icontains=query))
    clients = list(qs.order_by(*order_by))

    recent_blasts = list(
        MarketingBlast.objects.annotate(
            recipients_count=Count("logs"),
            success_count=Count("logs", filter=Q(logs__success=True)),
        )
        .prefetch_related(
            Prefetch(
                "logs",
                queryset=SmsLog.objects.filter(kind=SmsKind.MARKETING)
                .select_related("client")
                .order_by("id"),
            )
        )
        .order_by("-id")[:10]
    )

    return render(
        request,
        "workshop/marketing_sms.html",
        {
            "clients": clients,
            "cfg": cfg,
            "default_text": cfg.marketing_default_text,
            "query": query,
            "sort": sort,
            "recent_blasts": recent_blasts,
        },
    )


@require_POST
def marketing_blast_delete(request: HttpRequest, blast_id: int):
    from workshop.models import MarketingBlast

    blast = get_object_or_404(MarketingBlast, pk=blast_id)
    details = f"#{blast.id} recipients={blast.logs.count()}"
    blast.delete()
    log_action(request, "marketing_blast_delete", entity_type="messaging", entity_id=blast_id, details=details)
    messages.success(request, "Рассылка удалена из очереди")
    return redirect("marketing_sms")


@require_POST
def marketing_message_delete(request: HttpRequest, log_id: int):
    """Совместимость: удаление одиночного лога или всей рассылки, если лог к ней привязан."""
    from workshop.models import SmsKind, SmsLog

    log = get_object_or_404(SmsLog, pk=log_id, kind=SmsKind.MARKETING)
    if log.blast_id:
        blast = log.blast
        details = f"blast=#{blast.id} via log=#{log.id}"
        blast.delete()
        log_action(request, "marketing_blast_delete", entity_type="messaging", entity_id=blast.id, details=details)
        messages.success(request, "Рассылка удалена из очереди")
    else:
        details = f"#{log.id} {log.phone}"
        log.delete()
        log_action(request, "marketing_log_delete", entity_type="messaging", entity_id=log_id, details=details)
        messages.success(request, "Сообщение удалено из очереди")
    return redirect("marketing_sms")


@require_GET
def max_bot_qr(request: HttpRequest):
    """PNG QR-код со ссылкой на бота Max из настроек."""
    from io import BytesIO

    from workshop.models import SmsSettings

    try:
        import qrcode
    except ImportError:
        return HttpResponse("Установите qrcode: pip install qrcode", status=500, content_type="text/plain")

    cfg = SmsSettings.get_solo()
    link = (cfg.bot_link or "").strip()
    if not link and cfg.bot_username:
        link = f"https://max.ru/{cfg.bot_username.lstrip('@')}"
    if not link:
        return HttpResponse("Ссылка на бота не задана", status=404, content_type="text/plain")

    qr = qrcode.QRCode(version=None, box_size=8, border=2)
    qr.add_data(link)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = BytesIO()
    img.save(buf, format="PNG")
    response = HttpResponse(buf.getvalue(), content_type="image/png")
    response["Cache-Control"] = "no-store"
    return response


@csrf_exempt
@require_http_methods(["GET", "POST"])
def max_webhook(request: HttpRequest):
    """Webhook endpoint for Max (if public HTTPS is available)."""
    from workshop.messaging import process_updates_payload

    if request.method == "GET":
        return HttpResponse("max webhook ok")
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return HttpResponse("bad json", status=400)
    process_updates_payload(payload)
    return HttpResponse("ok")


def audit_log_list(request: HttpRequest):
    q = request.GET.get("q", "").strip()
    logs = AuditLog.objects.all()
    if q:
        logs = logs.filter(
            Q(username__icontains=q)
            | Q(action__icontains=q)
            | Q(details__icontains=q)
            | Q(entity_id__icontains=q)
        )
    return render(request, "workshop/audit_log.html", {"logs": logs[:300], "query": q})


@require_GET
def audit_log_export(request: HttpRequest):
    q = request.GET.get("q", "").strip()
    logs = AuditLog.objects.all().order_by("-id")
    if q:
        logs = logs.filter(
            Q(username__icontains=q)
            | Q(action__icontains=q)
            | Q(details__icontains=q)
            | Q(entity_id__icontains=q)
        )
    lines: list[str] = []
    for log in logs[:20000]:
        when = timezone.localtime(log.created_at).strftime("%Y-%m-%d %H:%M:%S") if log.created_at else ""
        lines.append(
            "\t".join(
                [
                    when,
                    log.username or "",
                    log.action or "",
                    log.entity_type or "",
                    str(log.entity_id or ""),
                    (log.details or "").replace("\n", " ").replace("\r", " "),
                    log.ip_address or "-",
                ]
            )
        )
    stamp = timezone.localdate().isoformat()
    filename = f"log+{stamp}.log"
    content = "when\tuser\taction\tentity\tid\tdetails\tip\n" + "\n".join(lines) + ("\n" if lines else "")
    response = HttpResponse(content, content_type="text/plain; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    log_action(request, "audit_log_export", entity_type="audit", details=filename)
    return response


def order_print(request: HttpRequest, order_id: int):
    order = get_object_or_404(Order.objects.select_related("client"), pk=order_id)
    log_action(
        request,
        "order_print_view",
        entity_type="order",
        entity_id=order.id,
        details=order.order_number,
    )
    return render(
        request,
        "workshop/print_order.html",
        {
            "order": order,
            "lines": order.lines.all(),
            "company_name": settings.COMPANY_NAME,
            "company_phone": settings.COMPANY_PHONE,
            "quality_phone": settings.QUALITY_PHONE,
            "company_address": settings.COMPANY_ADDRESS,
        },
    )


def order_pdf(request: HttpRequest, order_id: int):
    order = get_object_or_404(Order.objects.select_related("client"), pk=order_id)
    pdf_bytes = build_order_pdf(order, list(order.lines.all()))
    log_action(
        request,
        "order_pdf",
        entity_type="order",
        entity_id=order.id,
        details=order.order_number,
    )
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{order.order_number}.pdf"'
    return response


@require_POST
def order_print_direct(request: HttpRequest, order_id: int):
    order = get_object_or_404(Order.objects.select_related("client"), pk=order_id)
    pdf_bytes = build_order_pdf(order, list(order.lines.all()))
    username = str(request.session.get("workshop_username") or "")
    jobs = enqueue_pdf_print(
        pdf_bytes=pdf_bytes,
        title=order.order_number,
        doc_type="order",
        entity_type="order",
        entity_id=order.id,
        username=username,
        copies=PRINT_COPIES,
        request=request,
    )
    messages.success(
        request,
        f"Заказ-наряд добавлен в очередь печати ({PRINT_COPIES} экз., задания {[j.id for j in jobs]})",
    )
    return redirect("order_detail", order_id=order.id)


@require_http_methods(["GET", "POST"])
def acceptance_list_create(request: HttpRequest):
    if request.method == "POST":
        client_id = request.POST.get("client_id", "").strip()
        if not client_id:
            messages.warning(request, "Выберите клиента")
            return redirect("acceptance_acts")
        client = get_object_or_404(Client, pk=int(client_id))
        device_type = request.POST.get("device_type", DeviceType.PC)
        if device_type not in dict(DeviceType.choices):
            device_type = DeviceType.PC
        defect = request.POST.get("declared_defect", "").strip()
        if not defect:
            messages.warning(request, "Укажите заявленную неисправность")
            return redirect("acceptance_acts")
        order = None
        order_id = request.POST.get("order_id", "").strip()
        if order_id:
            order = Order.objects.filter(pk=int(order_id)).first()
        act = AcceptanceAct.objects.create(
            act_number=next_numbered("ACT", AcceptanceAct, "act_number"),
            client=client,
            order=order,
            device_type=device_type,
            brand_model=request.POST.get("brand_model", "").strip(),
            serial_number=request.POST.get("serial_number", "").strip(),
            accessories=request.POST.get("accessories", "").strip(),
            appearance=request.POST.get("appearance", "").strip(),
            declared_defect=defect,
            password_info=request.POST.get("password_info", "").strip(),
            notes=request.POST.get("notes", "").strip(),
        )
        messages.success(request, f"Акт {act.act_number} создан")
        log_action(request, "acceptance_create", entity_type="acceptance", entity_id=act.id, details=act.act_number)
        return redirect("acceptance_detail", act_id=act.id)

    return render(
        request,
        "workshop/acceptance_list.html",
        {
            "acts": AcceptanceAct.objects.select_related("client", "order")[:100],
            "clients": Client.objects.all(),
            "orders": Order.objects.select_related("client")[:100],
            "device_types": [c[0] for c in DeviceType.choices],
            "act_statuses": AcceptanceActStatus.choices,
        },
    )


def acceptance_detail(request: HttpRequest, act_id: int):
    act = get_object_or_404(AcceptanceAct.objects.select_related("client", "order"), pk=act_id)
    return render(
        request,
        "workshop/acceptance_detail.html",
        {
            "act": act,
            "act_statuses": AcceptanceActStatus.choices,
        },
    )


def acceptance_print(request: HttpRequest, act_id: int):
    act = get_object_or_404(AcceptanceAct.objects.select_related("client", "order"), pk=act_id)
    log_action(
        request,
        "acceptance_print_view",
        entity_type="acceptance",
        entity_id=act.id,
        details=act.act_number,
    )
    return render(
        request,
        "workshop/print_acceptance.html",
        {
            "act": act,
            "company_name": settings.COMPANY_NAME,
            "company_phone": settings.COMPANY_PHONE,
            "company_address": settings.COMPANY_ADDRESS,
            "master_sign": settings.MASTER_SIGN,
        },
    )


def acceptance_pdf(request: HttpRequest, act_id: int):
    act = get_object_or_404(AcceptanceAct.objects.select_related("client", "order"), pk=act_id)
    pdf_bytes = build_acceptance_act_pdf(act)
    log_action(
        request,
        "acceptance_pdf",
        entity_type="acceptance",
        entity_id=act.id,
        details=act.act_number,
    )
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{act.act_number}.pdf"'
    return response


@require_POST
def acceptance_print_direct(request: HttpRequest, act_id: int):
    act = get_object_or_404(AcceptanceAct.objects.select_related("client", "order"), pk=act_id)
    pdf_bytes = build_acceptance_act_pdf(act)
    username = str(request.session.get("workshop_username") or "")
    jobs = enqueue_pdf_print(
        pdf_bytes=pdf_bytes,
        title=act.act_number,
        doc_type="acceptance",
        entity_type="acceptance",
        entity_id=act.id,
        username=username,
        copies=PRINT_COPIES,
        request=request,
    )
    messages.success(
        request,
        f"Акт добавлен в очередь печати ({PRINT_COPIES} экз., задания {[j.id for j in jobs]})",
    )
    return redirect("acceptance_detail", act_id=act.id)


@require_POST
def acceptance_delete(request: HttpRequest, act_id: int):
    act = get_object_or_404(AcceptanceAct, pk=act_id)
    details = act.act_number
    act.delete()
    log_action(request, "acceptance_delete", entity_type="acceptance", entity_id=act_id, details=details)
    messages.success(request, "Акт удалён")
    return redirect("acceptance_acts")

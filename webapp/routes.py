from __future__ import annotations

from io import BytesIO
import os
import platform
import subprocess
import tempfile

from flask import Flask, flash, g, redirect, render_template, request, send_file, url_for

from database import Database
from webapp.utils import normalize_rf_phone


def _excel_available() -> bool:
    try:
        import openpyxl  # noqa: F401

        return True
    except ImportError:
        return False


def _build_clients_workbook(rows: list[tuple[str, str, str]]) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Клиенты"
    ws.append(["номер телефона", "имя", "комментарий"])
    for phone, name, comment in rows:
        ws.append([phone, name, comment])
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _parse_clients_workbook(file_storage) -> list[tuple[str, str, str]]:
    from openpyxl import load_workbook

    wb = load_workbook(file_storage, read_only=True, data_only=True)
    ws = wb.active
    rows: list[tuple[str, str, str]] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row:
            continue
        values = list(row) + [None, None, None]
        phone = "" if values[0] is None else str(values[0]).strip()
        name = "" if values[1] is None else str(values[1]).strip()
        comment = "" if values[2] is None else str(values[2]).strip()
        if idx == 1 and ("телефон" in phone.lower() or phone.lower() in {"phone", "номер телефона"}):
            continue
        rows.append((phone, name, comment))
    return rows


def register_routes(app: Flask) -> None:
    def get_db() -> Database:
        db = g.get("db")
        if db is None:
            db = Database()
            g.db = db
        return db

    @app.teardown_appcontext
    def close_db(_exception):
        db = g.pop("db", None)
        if db is not None:
            db.close()

    def build_order_pdf(order, services) -> bytes:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.pdfgen import canvas
        except ImportError:
            raise RuntimeError("Для экспорта PDF установите reportlab: pip install reportlab")

        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        y = height - 40

        font_name = "Helvetica"
        font_candidates = [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/dejavu/DejaVuSans.ttf",
            "C:/Windows/Fonts/arial.ttf",
            "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        ]
        for candidate in font_candidates:
            if os.path.exists(candidate):
                try:
                    pdfmetrics.registerFont(TTFont("AppFont", candidate))
                    font_name = "AppFont"
                    break
                except Exception:
                    continue

        c.setFont(font_name, 16)
        c.drawString(40, y, f"Заказ-наряд {order['order_number']}")
        y -= 22
        c.setFont(font_name, 11)
        c.drawString(40, y, "ИТ- Мастерская, тел.: +7 (918) 802 - 87 - 67")
        y -= 16
        c.drawString(40, y, "Контроль качества: +7 (962) 550 - 78 - 32")
        y -= 16
        c.drawString(40, y, "Адрес: р. Татарстан, д.Куюки, ул. 24 квартал дом 1")
        y -= 16
        c.drawString(40, y, f"Клиент: {order['client_name']}  {order['phone']}")
        y -= 16
        c.drawString(40, y, f"Дата: {order['created_date']}")
        y -= 22

        c.drawString(40, y, f"Устройство: {order['device_type']}")
        y -= 16
        c.drawString(40, y, f"Доп. периферия: {order['extra_periphery'] or '-'}")
        y -= 16

        c.setFont(font_name, 10)
        c.drawString(40, y, "Услуга")
        c.drawString(350, y, "Цена")
        c.drawString(430, y, "Кол-во")
        c.drawString(500, y, "Сумма")
        y -= 10
        c.line(40, y, width - 40, y)
        y -= 14

        for service in services:
            if y < 70:
                c.showPage()
                c.setFont(font_name, 10)
                y = height - 50
            line_total = float(service["price"]) * int(service["quantity"])
            c.drawString(40, y, str(service["service_name"])[:52])
            c.drawRightString(400, y, f"{float(service['price']):.2f}")
            c.drawRightString(470, y, f"{int(service['quantity'])}")
            c.drawRightString(555, y, f"{line_total:.2f}")
            y -= 14

        y -= 8
        c.line(40, y, width - 40, y)
        y -= 20
        c.setFont(font_name, 12)
        c.drawRightString(width - 40, y, f"ИТОГО: {float(order['total_sum']):.2f}")
        y -= 24
        c.setFont(font_name, 10)
        c.drawString(
            40,
            y,
            "Гарантия: На выполненные работы и установленные новые детали предоставляется гарантия 3 месяца.",
        )
        y -= 14
        c.drawString(
            40,
            y,
            "Гарантия не распространяется на программное обеспечение и устранение последствий некорректного использования.",
        )
        y -= 16
        c.drawString(40, y, f"Техническая информация/рекомендации: {order['technical_notes'] or '-'}")
        y -= 20
        c.drawString(40, y, "Исполнитель: _________________ / Григорьев Д.В")
        y -= 14
        c.drawString(40, y, "(Подпись) (Ф.И.О.)")
        y -= 18
        c.drawString(40, y, "Заказчик с работами ознакомлен, результат меня устраивает, претензий не имею.")
        y -= 16
        c.drawString(40, y, "Заказчик:___________________ / __________________________ / «        » _______ 2026г.")
        y -= 14
        c.drawString(40, y, "(Подпись) (Ф.И.О.) (Дата)")
        c.save()

        buffer.seek(0)
        return buffer.getvalue()

    @app.route("/")
    def dashboard():
        from webapp.network import get_lan_ipv4_addresses

        db = get_db()
        orders = db.get_all_orders()
        port = int(os.getenv("IT_MASTER_PORT", "8000"))
        lan_urls = [f"http://{ip}:{port}" for ip in get_lan_ipv4_addresses() if ip != "127.0.0.1"]
        return render_template("dashboard.html", recent_orders=orders[:20], lan_urls=lan_urls)

    @app.route("/statistics")
    def statistics():
        db = get_db()
        period = request.args.get("period", "month")
        if period not in {"week", "month", "year"}:
            period = "month"
        stats = db.get_statistics(period)
        return render_template(
            "statistics.html",
            period=period,
            stats=stats,
            monthly=db.get_monthly_statistics(12),
            top_clients=db.get_top_clients_by_spent(10),
        )

    @app.route("/clients", methods=["GET", "POST"])
    def clients():
        db = get_db()
        if request.method == "POST":
            name = request.form.get("name", "").strip()
            phone_raw = request.form.get("phone", "").strip()
            comment = request.form.get("comment", "").strip()
            normalized = normalize_rf_phone(phone_raw)
            if not name:
                flash("Введите имя клиента", "warning")
                return redirect(url_for("clients"))
            if not normalized:
                flash("Введите корректный номер РФ (например, +7 962 550 7832)", "warning")
                return redirect(url_for("clients"))
            if db.find_client_by_phone(normalized):
                flash("Клиент с таким телефоном уже существует", "warning")
                return redirect(url_for("clients"))
            client_id = db.create_client(name, normalized, comment)
            flash(
                "Клиент успешно добавлен" if client_id else "Не удалось создать клиента",
                "success" if client_id else "danger",
            )
            return redirect(url_for("clients"))

        query = request.args.get("q", "").strip()
        rows = db.search_clients(query) if query else db.get_all_clients()
        return render_template("clients.html", clients=rows, query=query)

    @app.route("/clients/export.xlsx")
    def export_clients_excel():
        if not _excel_available():
            flash("Для Excel установите openpyxl: pip install openpyxl", "warning")
            return redirect(url_for("clients"))
        db = get_db()
        payload = _build_clients_workbook(db.export_clients_excel_rows())
        return send_file(
            BytesIO(payload),
            as_attachment=True,
            download_name="clients.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.route("/clients/import", methods=["POST"])
    def import_clients_excel():
        if not _excel_available():
            flash("Для Excel установите openpyxl: pip install openpyxl", "warning")
            return redirect(url_for("clients"))
        upload = request.files.get("excel_file")
        if not upload or not upload.filename:
            flash("Выберите Excel-файл", "warning")
            return redirect(url_for("clients"))
        try:
            rows = _parse_clients_workbook(upload)
            result = get_db().import_clients_from_rows(rows)
            flash(
                f"Импорт завершён: добавлено {result['imported']}, "
                f"пропущено существующих {result['skipped_existing']}, "
                f"некорректных строк {result['skipped_invalid']}",
                "success",
            )
        except Exception as err:
            flash(f"Не удалось импортировать Excel: {err}", "danger")
        return redirect(url_for("clients"))

    @app.route("/services", methods=["GET", "POST"])
    def services():
        db = get_db()
        if request.method == "POST":
            name = request.form.get("name", "").strip()
            price_raw = request.form.get("price", "").strip().replace(",", ".")
            category = request.form.get("category", "").strip() or "Основные"
            if not name:
                flash("Введите название услуги", "warning")
                return redirect(url_for("services"))
            try:
                price = float(price_raw)
            except ValueError:
                flash("Введите корректную цену", "warning")
                return redirect(url_for("services"))
            if price < 0:
                flash("Цена не может быть отрицательной", "warning")
                return redirect(url_for("services"))
            if db.get_service_by_name(name):
                flash("Услуга с таким названием уже существует", "warning")
                return redirect(url_for("services"))
            service_id = db.add_service(name, price, category)
            flash(
                "Услуга добавлена" if service_id else "Не удалось добавить услугу",
                "success" if service_id else "danger",
            )
            return redirect(url_for("services"))

        return render_template(
            "services.html",
            services=db.get_all_services(),
            categories=db.get_categories(),
            catalog_tree=db.get_service_catalog_tree(active_only=False),
        )

    @app.route("/orders")
    def orders():
        return render_template("orders.html", orders=get_db().get_all_orders())

    @app.route("/orders/new", methods=["GET", "POST"])
    def create_order():
        db = get_db()
        if request.method == "POST":
            client_id_raw = request.form.get("client_id", "").strip()
            client_id = int(client_id_raw) if client_id_raw else None
            order_id, _ = db.create_order(client_id)
            flash("Заказ создан", "success")
            return redirect(url_for("order_detail", order_id=order_id))
        return render_template("order_new.html", clients=db.get_all_clients())

    @app.route("/orders/<int:order_id>")
    def order_detail(order_id: int):
        db = get_db()
        order = db.get_order_by_id(order_id)
        if not order:
            flash("Заказ не найден", "warning")
            return redirect(url_for("orders"))
        return render_template(
            "order_detail.html",
            order=order,
            services=db.get_order_services(order_id),
            catalog=db.get_active_services(),
            catalog_tree=db.get_service_catalog_tree(active_only=True),
            device_types=["ПК", "Ноутбук", "Телефон", "Телевизор"],
        )

    @app.route("/orders/<int:order_id>/meta", methods=["POST"])
    def update_order_meta(order_id: int):
        db = get_db()
        order = db.get_order_by_id(order_id)
        if not order:
            flash("Заказ не найден", "warning")
            return redirect(url_for("orders"))
        device_type = request.form.get("device_type", "ПК").strip() or "ПК"
        extra_periphery = request.form.get("extra_periphery", "").strip()
        technical_notes = request.form.get("technical_notes", "").strip()
        db.update_order_meta(order_id, device_type, extra_periphery, technical_notes)
        flash("Данные заказ-наряда сохранены", "success")
        return redirect(url_for("order_detail", order_id=order_id))

    @app.route("/orders/<int:order_id>/print")
    def print_order(order_id: int):
        db = get_db()
        order = db.get_order_by_id(order_id)
        if not order:
            flash("Заказ не найден", "warning")
            return redirect(url_for("orders"))
        return render_template(
            "print_order.html",
            order=order,
            services=db.get_order_services(order_id),
            company_phone="+7 (918) 802 - 87 - 67",
            quality_phone="+7 (962) 550 - 78 - 32",
            company_name="ИТ- Мастерская",
            company_address="р. Татарстан, д.Куюки, ул. 24 квартал дом 1",
        )

    @app.route("/orders/<int:order_id>/pdf")
    def order_pdf(order_id: int):
        db = get_db()
        order = db.get_order_by_id(order_id)
        if not order:
            flash("Заказ не найден", "warning")
            return redirect(url_for("orders"))
        try:
            pdf_bytes = build_order_pdf(order, db.get_order_services(order_id))
        except RuntimeError as err:
            flash(str(err), "warning")
            return redirect(url_for("order_detail", order_id=order_id))
        return send_file(
            BytesIO(pdf_bytes),
            as_attachment=True,
            download_name=f"{order['order_number']}.pdf",
            mimetype="application/pdf",
        )

    @app.route("/orders/<int:order_id>/print-direct", methods=["POST"])
    def print_order_direct(order_id: int):
        db = get_db()
        order = db.get_order_by_id(order_id)
        if not order:
            flash("Заказ не найден", "warning")
            return redirect(url_for("orders"))
        try:
            pdf_bytes = build_order_pdf(order, db.get_order_services(order_id))
        except RuntimeError as err:
            flash(str(err), "warning")
            return redirect(url_for("order_detail", order_id=order_id))

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        try:
            tmp.write(pdf_bytes)
            tmp.flush()
            tmp.close()
            if platform.system() == "Windows":
                os.startfile(tmp.name, "print")
            else:
                subprocess.run(["lp", tmp.name], check=True)
                try:
                    os.unlink(tmp.name)
                except OSError:
                    pass
            flash("Документ отправлен на принтер", "success")
        except Exception as err:
            flash(f"Не удалось отправить документ на принтер: {err}", "danger")
            try:
                os.unlink(tmp.name)
            except OSError:
                pass
        return redirect(url_for("order_detail", order_id=order_id))

    @app.route("/orders/<int:order_id>/add-service", methods=["POST"])
    def add_order_service(order_id: int):
        db = get_db()
        order = db.get_order_by_id(order_id)
        if not order:
            flash("Заказ не найден", "warning")
            return redirect(url_for("orders"))
        service_name = request.form.get("service_name", "").strip()
        try:
            qty = max(1, int(request.form.get("quantity", "1").strip()))
        except ValueError:
            qty = 1
        service = db.get_service_by_name(service_name)
        if not service:
            flash("Услуга не найдена", "warning")
            return redirect(url_for("order_detail", order_id=order_id))
        db.add_service_to_order(order_id, service_name, float(service["price"]), qty)
        db.update_order_total(order_id)
        flash("Услуга добавлена в заказ", "success")
        return redirect(url_for("order_detail", order_id=order_id))

    @app.route("/orders/<int:order_id>/line/<int:line_id>/delete", methods=["POST"])
    def delete_order_line(order_id: int, line_id: int):
        db = get_db()
        db.cursor.execute("DELETE FROM order_service_lines WHERE id = ? AND order_id = ?", (line_id, order_id))
        db.cursor.execute("DELETE FROM order_services WHERE id = ? AND order_id = ?", (line_id, order_id))
        db.conn.commit()
        db.update_order_total(order_id)
        flash("Строка удалена", "success")
        return redirect(url_for("order_detail", order_id=order_id))

    @app.route("/orders/<int:order_id>/delete", methods=["POST"])
    def delete_order(order_id: int):
        get_db().delete_order(order_id)
        flash("Заказ удалён", "success")
        return redirect(url_for("orders"))
